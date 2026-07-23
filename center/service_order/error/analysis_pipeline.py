from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Callable, Iterable
import json
import re
import sqlite3
import uuid

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from service_order.error.pattern_matching import (
    character_bigram_counter,
    compact_signature,
    counter_cosine,
    informative_token_count,
    ordered_distinct_token_match,
)
from service_order.error.privacy import mask_personal_data_frame


BASE_DIR = Path(__file__).resolve().parent
JSON_DIR = BASE_DIR / "json"
SAP_ID_PATH = JSON_DIR / "sap_id.json"
EXCEPT_LIST_PATH = JSON_DIR / "except_list.json"
NORMAL_RULES_PATH = JSON_DIR / "normal_rules.json"
ERROR_RULES_PATH = JSON_DIR / "error_rules.json"

# Temporary test/migration hooks. Production inference and learning use only
# NORMAL_RULES_PATH and ERROR_RULES_PATH. These names remain so older unit
# fixtures can exercise the legacy classifier without requiring legacy files
# in the runtime json directory.
_LEGACY_KEYWORD_PATH = JSON_DIR / "keyword.json"
_LEGACY_CONFLICT_KEYWORD_PATH = JSON_DIR / "conflict_keyword.json"
_LEGACY_PRIORITY_KEYWORD_PATH = JSON_DIR / "priority_keyword.json"
_LEGACY_ERROR_PATTERN_PATH = JSON_DIR / "error_pattern.json"
_LEGACY_AUTO_NORMAL_PATTERN_PATH = JSON_DIR / "auto_normal_pattern.json"
_LEGACY_SCOPED_NORMAL_PATTERN_PATH = JSON_DIR / "scoped_normal_pattern.json"
KEYWORD_PATH = _LEGACY_KEYWORD_PATH
CONFLICT_KEYWORD_PATH = _LEGACY_CONFLICT_KEYWORD_PATH
PRIORITY_KEYWORD_PATH = _LEGACY_PRIORITY_KEYWORD_PATH
ERROR_PATTERN_PATH = _LEGACY_ERROR_PATTERN_PATH
AUTO_NORMAL_PATTERN_PATH = _LEGACY_AUTO_NORMAL_PATTERN_PATH
SCOPED_NORMAL_PATTERN_PATH = _LEGACY_SCOPED_NORMAL_PATTERN_PATH
ERROR_LEARNING_BASELINE_PATH = (
    BASE_DIR.parent / "data" / "error_learning_baseline.sqlite3"
)

ERROR_SIMILARITY_MIN_SCORE = 0.75
ERROR_SIMILARITY_MAX_NORMAL_SCORE = 0.55
ERROR_SIMILARITY_MIN_MARGIN = 0.20
ERROR_SIMILARITY_MIN_INFORMATIVE_TOKENS = 2
ERROR_SIMILARITY_PREFILTER_LIMIT = 30

ACTIVE_ERROR_PATTERN_STATUSES = frozenset(
    {
        "active",
        "auto_error",
        "confirmed",
        "confirmed_error",
        "enabled",
        "오생성",
        "확정",
        "확정오생성",
    }
)
REVIEW_ERROR_PATTERN_STATUSES = frozenset(
    {
        "ambiguous",
        "ambiguous_review",
        "review",
        "검토",
    }
)
INACTIVE_ERROR_PATTERN_STATUSES = frozenset(
    {
        "disabled",
        "inactive",
        "normal",
        "retired",
        "비활성",
    }
)

BUSINESS_BY_CENTER = {
    "H071": "중부",
    "H072": "북부",
    "H073": "남부",
    "H074": "동부",
    "H075": "서부",
}

BUSINESS_CREATOR_PREFIXES = frozenset(
    {"CSC71", "CSC72", "CSC73", "CSC74", "CSC75"}
)
BUSINESS_BY_CREATOR_PREFIX = {
    "CSC71": "중부",
    "CSC72": "북부",
    "CSC73": "남부",
    "CSC74": "동부",
    "CSC75": "서부",
}
CREATOR_BUSINESS_EXCEPTIONS = {"CSC7013": "중부"}
SERVICE_QUALITY_CREATOR_PREFIX = "CSC70"

# 관리자 페이지 결과는 전처리된 단일 ``서비스처리센터`` 열을 사용합니다.
# 전처리 파일에는 ``내역2``가 없고, 후보군 파일을 만들 때만 ``내역``
# 바로 뒤에 추가합니다.
CANDIDATE_COLUMN_ORDER = (
    "서비스처리센터",
    "고객방문일",
    "고객방문시간",
    "우선순위",
    "오더번호",
    "상태",
    "대분류",
    "중분류",
    "소분류",
    "내역",
    "내역2",
    "세대",
    "고객번호",
    "계약계정번호",
    "정보",
    "정보(BP)",
    "최초공급일",
    "주소",
    "구주소",
    "동분리",
    "건물동",
    "호수",
    "전입/전출일",
    "오더생성일",
    "오더생성시간",
    "오더생성자",
    "생성인",
    "생성부서",
    "사업부",
    "부서확인일",
    "부서확인시간",
    "부서확인자",
    "기사배정일",
    "기사배정시간",
    "배정된기사",
    "마감일",
    "마감시간",
    "마감자",
    "변경일",
    "변경시간",
    "변경자",
    "수수료금액",
    "영수증번호",
)
PREPROCESSED_COLUMN_ORDER = tuple(
    column for column in CANDIDATE_COLUMN_ORDER if column != "내역2"
)

HEADER_SIGNATURE = {
    "오더번호",
    "소분류",
    "내역",
    "오더생성일",
    "오더생성자",
}
MAX_HEADER_SCAN_ROWS = 30

PERSON_NAME_COLUMNS = ["생성인", "부서확인자", "배정된기사", "마감자"]
KOREAN_SURNAMES = set(
    "김이박최정강조윤장임한오서신권황안송전홍유고문양손배백허남심노하곽성차주우구민류나진지엄채원천방공현함변염여추도소석선설마길연위표명기반왕금옥육인맹제모탁국어은편용예봉사부가복태목형피두감음빈동온호범좌팽승간상시갈단"
)
PERSON_SUFFIXES = (
    "매니저님",
    "기사님",
    "과장님",
    "대리님",
    "사원님",
    "매니저",
    "기사",
    "과장",
    "대리",
    "사원",
    "님",
)
NAME_CONTEXT_TOKENS = {
    "배정",
    "배정요청",
    "매니저",
    "기사",
    "담당",
    "담당자",
    "확인자",
    "접수자",
}
NON_PERSON_NAME_BASES = {
    "안전",
    "민원",
    "고객",
    "담당",
    "현장",
    "검침",
    "요금",
    "전입",
    "전출",
    "공급",
    "서비스",
    "시설",
    "설치",
    "송달",
    "청구서",
    "고지서",
    "계량기",
    "타이머",
    "자원",
}

TOKEN_PATTERN = re.compile(r"(?u)\b[가-힣a-z][가-힣a-z]+\b")


def clean_text(series: pd.Series) -> pd.Series:
    return series.astype("string").str.strip()


ADDRESS_AREA_PATTERN = re.compile(
    r"(?:^|[\s(])([가-힣]+(?:동|읍|면|가))(?=[\s)]|$)"
)


def split_address(value: object) -> tuple[object, object, object]:
    if pd.isna(value):
        return pd.NA, pd.NA, pd.NA
    combined = str(value).strip()
    if not combined:
        return pd.NA, pd.NA, pd.NA

    road_address = combined
    old_address = combined
    opening = combined.find("(")
    if opening > 0 and combined.endswith(")"):
        road_address = combined[:opening].strip()
        old_address = combined[opening + 1 : -1].strip()

    area_match = ADDRESS_AREA_PATTERN.search(old_address)
    if area_match is None:
        area_match = ADDRESS_AREA_PATTERN.search(road_address)
    area = area_match.group(1) if area_match else pd.NA
    return road_address or pd.NA, old_address or pd.NA, area


def normalize_address_columns(data: pd.DataFrame) -> pd.DataFrame:
    if "주소" not in data.columns:
        data["주소"] = pd.NA
    combined_address = data["주소"].copy()
    derived = combined_address.map(split_address)
    road_address = derived.map(lambda values: values[0])
    old_address = derived.map(lambda values: values[1])
    area = derived.map(lambda values: values[2])

    # Compact Amaranth exports contain one combined address column. Monthly
    # candidate files use three columns, so split only when the companion
    # columns are absent; already-preprocessed inputs retain their values.
    if "구주소" not in data.columns:
        data["주소"] = road_address
        data["구주소"] = old_address
    else:
        data["구주소"] = clean_text(data["구주소"])
    if "동분리" not in data.columns:
        data["동분리"] = area
    else:
        data["동분리"] = clean_text(data["동분리"])
    return data


def detect_excel_header_row(path: Path) -> tuple[int, str]:
    """Return the zero-based header row and first worksheet name.

    Amaranth exports can contain a report title, print date, blank rows, and
    blank leading columns before the actual service-order header. Standard
    one-row-header workbooks are supported by the same detection.
    """

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            raise ValueError("Excel 파일에 읽을 수 있는 시트가 없습니다.")
        worksheet = workbook.worksheets[0]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS),
            start=1,
        ):
            values = {
                str(cell.value).strip()
                for cell in row
                if cell.value is not None and str(cell.value).strip()
            }
            if HEADER_SIGNATURE.issubset(values):
                return row_number - 1, worksheet.title
    finally:
        workbook.close()

    expected = ", ".join(sorted(HEADER_SIGNATURE))
    raise ValueError(
        f"서비스오더 열 이름 행을 찾지 못했습니다. 필요한 열: {expected}"
    )


def read_order_excel(path: Path) -> tuple[pd.DataFrame, dict[str, object]]:
    header_row, sheet_name = detect_excel_header_row(path)
    raw = pd.read_excel(path, sheet_name=0, header=header_row)
    raw.columns = raw.columns.astype(str).str.strip()

    # Report decoration columns have no header and may contain a trailing '*'.
    # They are never service-order data columns.
    named_columns = [
        column
        for column in raw.columns
        if column and not column.lower().startswith("unnamed:")
    ]
    raw = raw.loc[:, named_columns]

    # Amaranth stores many values as strings padded with hundreds of spaces.
    # Trim every textual cell so the downloaded result matches the monthly
    # candidate files and so the report footer becomes an empty row.
    for column in raw.columns:
        if raw[column].dtype == object or isinstance(raw[column].dtype, pd.StringDtype):
            raw[column] = raw[column].map(
                lambda value: value.strip() if isinstance(value, str) else value
            )
            raw[column] = raw[column].replace("", pd.NA)
    raw = raw.dropna(axis=0, how="all")
    if "오더번호" in raw.columns:
        raw = raw.loc[raw["오더번호"].notna()]
    raw = raw.reset_index(drop=True)

    return raw, {
        "원본시트": sheet_name,
        "인식헤더행": header_row + 1,
    }


def find_date_columns(columns: Iterable[str]) -> list[str]:
    return [
        column
        for column in columns
        if str(column).endswith("일")
        or str(column).endswith("일자")
        or str(column).endswith("날짜")
    ]


def reorder_columns(data: pd.DataFrame) -> pd.DataFrame:
    ordered = data.copy()
    for column in PREPROCESSED_COLUMN_ORDER:
        if column not in ordered.columns:
            ordered[column] = pd.NA
    return ordered.loc[:, PREPROCESSED_COLUMN_ORDER]


def load_configuration() -> tuple[dict[str, str], set[str]]:
    for path in (
        SAP_ID_PATH,
        EXCEPT_LIST_PATH,
        NORMAL_RULES_PATH,
        ERROR_RULES_PATH,
    ):
        if not path.is_file():
            raise FileNotFoundError(f"설정 파일을 찾을 수 없습니다: {path}")

    with SAP_ID_PATH.open(encoding="utf-8-sig") as file:
        sap_id_map = json.load(file)
    with EXCEPT_LIST_PATH.open(encoding="utf-8-sig") as file:
        except_list = json.load(file)
    if not isinstance(except_list, list):
        raise ValueError("except_list.json의 최상위 값은 배열이어야 합니다.")
    normalized_sap_ids = {
        str(sap_id).strip().upper(): str(name).strip()
        for sap_id, name in sap_id_map.items()
    }
    normalized_except_keys = {
        re.sub(r"\s+", "", str(value).strip())
        for value in except_list
        if str(value).strip()
    }
    return normalized_sap_ids, normalized_except_keys


def preprocess_orders(
    raw: pd.DataFrame,
    sap_id_map: dict[str, str],
    except_keys: set[str],
    *,
    return_dashboard_totals: bool = False,
) -> (
    tuple[pd.DataFrame, dict[str, object]]
    | tuple[pd.DataFrame, pd.DataFrame, dict[str, object]]
):
    data = raw.copy()
    data.columns = data.columns.astype(str).str.strip()

    required = {
        "오더번호",
        "오더생성일",
        "오더생성자",
        "상태",
        "소분류",
        "내역",
    }
    missing = sorted(required - set(data.columns))
    if missing:
        raise KeyError(f"필수 열이 없습니다: {missing}")

    center_columns = {"고객서비스처리센터", "서비스처리센터"}
    if not center_columns.intersection(data.columns):
        raise KeyError(
            "필수 열이 없습니다: 고객서비스처리센터 또는 서비스처리센터"
        )

    # 원본 서비스처리센터는 사용하지 않습니다. 고객서비스처리센터를
    # 유일한 서비스처리센터 열로 바꾸고 H051을 H073으로 병합합니다.
    if "고객서비스처리센터" in data.columns:
        normalized_center = clean_text(data["고객서비스처리센터"]).str.upper()
    else:
        # 이미 전처리된 단일 센터 열 파일도 다시 분석할 수 있게 합니다.
        normalized_center = clean_text(data["서비스처리센터"]).str.upper()
    data["서비스처리센터"] = normalized_center.replace({"H051": "H073"})
    if "고객서비스처리센터" in data.columns:
        data = data.drop(columns="고객서비스처리센터")
    data["오더생성자"] = clean_text(data["오더생성자"])
    data["상태"] = clean_text(data["상태"])
    data["소분류"] = clean_text(data["소분류"])
    data["내역"] = clean_text(data["내역"])
    data = normalize_address_columns(data)

    date_columns = find_date_columns(data.columns)
    invalid_date_counts: dict[str, int] = {}
    for column in date_columns:
        original_nonblank = data[column].notna() & clean_text(data[column]).ne("")
        converted = pd.to_datetime(data[column], errors="coerce").dt.normalize()
        invalid_date_counts[column] = int((original_nonblank & converted.isna()).sum())
        data[column] = converted

    creator_id = data["오더생성자"].str.upper()
    data["생성인"] = creator_id.map(sap_id_map).astype("string")

    department_inferred = "생성부서" not in data.columns
    if department_inferred:
        # The compact Amaranth service-order export omits 생성부서. The SAP
        # CSC71~CSC75 are the five business-department creator groups used by
        # the existing monthly candidate data. CSC70 is 서비스품질팀, except
        # CSC7013 (이두희), which is explicitly assigned to 중부 사업부.
        data["생성부서"] = pd.Series(pd.NA, index=data.index, dtype="string")
        inferred_business_mask = creator_id.str[:5].isin(BUSINESS_CREATOR_PREFIXES)
        inferred_business_mask |= creator_id.isin(CREATOR_BUSINESS_EXCEPTIONS)
        service_quality_mask = creator_id.str[:5].eq(SERVICE_QUALITY_CREATOR_PREFIX)
        service_quality_mask &= ~creator_id.isin(CREATOR_BUSINESS_EXCEPTIONS)
        data.loc[service_quality_mask, "생성부서"] = "서비스품질팀"
        data.loc[inferred_business_mask, "생성부서"] = "사업부"
    else:
        data["생성부서"] = clean_text(data["생성부서"])
        inferred_business_mask = pd.Series(False, index=data.index)
        service_quality_mask = pd.Series(False, index=data.index)

    creator_business = creator_id.str[:5].map(BUSINESS_BY_CREATOR_PREFIX)
    for creator, business in CREATOR_BUSINESS_EXCEPTIONS.items():
        creator_business = creator_business.mask(creator_id.eq(creator), business)
    if "사업부" in data.columns:
        existing_business = clean_text(data["사업부"])
        existing_business = existing_business.where(
            existing_business.isin(BUSINESS_BY_CENTER.values())
        )
        creator_business = creator_business.fillna(existing_business)
    center_business = data["서비스처리센터"].map(BUSINESS_BY_CENTER)
    data["사업부"] = creator_business.fillna(center_business).astype("string")

    business_department_mask = data["생성부서"].eq("사업부").fillna(False)
    category_keys = data["소분류"].map(
        lambda value: re.sub(r"\s+", "", str(value).strip())
    )
    excluded_mask = category_keys.isin(except_keys).fillna(False)
    cancelled_order_mask = data["상태"].eq("오더취소").fillna(False)
    dashboard_scope_mask = business_department_mask & ~cancelled_order_mask
    current_scope_mask = dashboard_scope_mask & ~excluded_mask
    result = data.loc[current_scope_mask].copy()
    result = reorder_columns(result).reset_index(drop=True)
    dashboard_totals = reorder_columns(
        data.loc[dashboard_scope_mask].copy()
    ).reset_index(drop=True)

    summary = {
        "원본행수": len(data),
        "결과행수": len(result),
        "표기집계행수": len(dashboard_totals),
        "표기집계추가행수": len(dashboard_totals) - len(result),
        "생성부서제외행수": int((~business_department_mask).sum()),
        "생성부서추론": department_inferred,
        "생성부서추론행수": int(
            (inferred_business_mask | service_quality_mask).sum()
        ),
        "사업부추론행수": int(inferred_business_mask.sum()),
        "서비스품질팀추론행수": int(service_quality_mask.sum()),
        "제외소분류행수": int(
            (business_department_mask & excluded_mask).sum()
        ),
        "오더취소제외행수": int(
            (business_department_mask & cancelled_order_mask).sum()
        ),
        "센터결측": int(result["서비스처리센터"].isna().sum()),
        "사업부결측": int(result["사업부"].isna().sum()),
        "생성인결측": int(result["생성인"].isna().sum()),
        "날짜변환실패": sum(invalid_date_counts.values()),
        "날짜열": date_columns,
    }
    if return_dashboard_totals:
        return result, dashboard_totals, summary
    return result, summary


def validate_preprocessed(data: pd.DataFrame) -> None:
    columns = list(data.columns)
    if not columns or columns[0] != "서비스처리센터":
        raise AssertionError("서비스처리센터가 첫 번째 열이 아닙니다.")
    if tuple(columns) != PREPROCESSED_COLUMN_ORDER:
        raise AssertionError("전처리 결과의 열 또는 열 순서가 기준 형식과 다릅니다.")
    if columns.index("생성인") != columns.index("오더생성자") + 1:
        raise AssertionError("생성인이 오더생성자 바로 뒤에 있지 않습니다.")
    if columns.index("생성부서") != columns.index("생성인") + 1:
        raise AssertionError("생성부서가 생성인 바로 뒤에 있지 않습니다.")
    if columns.index("사업부") != columns.index("생성부서") + 1:
        raise AssertionError("사업부가 생성부서 바로 뒤에 있지 않습니다.")
    if "고객서비스처리센터" in columns:
        raise AssertionError("고객서비스처리센터 열 이름이 남아 있습니다.")
    if data["서비스처리센터"].eq("H051").any():
        raise AssertionError("H051이 H073으로 병합되지 않았습니다.")
    if not data["생성부서"].eq("사업부").all():
        raise AssertionError("생성부서가 사업부가 아닌 행이 남아 있습니다.")
    if data["상태"].eq("오더취소").any():
        raise AssertionError("상태가 오더취소인 행이 남아 있습니다.")
    for column in find_date_columns(columns):
        if not pd.api.types.is_datetime64_any_dtype(data[column]):
            raise AssertionError(f"날짜 타입이 아닙니다: {column}")


def make_excel_compatible(data: pd.DataFrame) -> pd.DataFrame:
    compatible = data.copy()
    for column in compatible.columns:
        if isinstance(compatible[column].dtype, pd.StringDtype):
            compatible[column] = compatible[column].astype(object)
            compatible[column] = compatible[column].where(
                compatible[column].notna(), None
            )
        elif pd.api.types.is_datetime64_any_dtype(compatible[column]):
            compatible[column] = pd.to_datetime(
                compatible[column], errors="coerce"
            ).astype("datetime64[ns]")
        compatible[column] = compatible[column].map(
            lambda value: (
                f"'{value}"
                if isinstance(value, str)
                and value.lstrip().startswith(("=", "+", "-", "@"))
                else value
            )
        )
    return compatible


def save_formatted_workbook(
    sheets: Iterable[tuple[str, pd.DataFrame]],
    path: Path,
) -> None:
    prepared_sheets = [
        (
            str(sheet_name),
            make_excel_compatible(mask_personal_data_frame(data)[0]),
        )
        for sheet_name, data in sheets
    ]
    if not prepared_sheets:
        raise ValueError("저장할 Excel 시트가 없습니다.")
    sheet_names = [sheet_name for sheet_name, _ in prepared_sheets]
    if len(sheet_names) != len(set(sheet_names)):
        raise ValueError("Excel 시트 이름이 중복되었습니다.")

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_name(
        f".{path.stem}_{uuid.uuid4().hex[:8]}_writing.xlsx"
    )
    try:
        with pd.ExcelWriter(temporary_path, engine="openpyxl") as writer:
            for sheet_name, excel_data in prepared_sheets:
                excel_data.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )
        workbook = load_workbook(temporary_path)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        body_font = Font(name="맑은 고딕", size=11)
        for sheet_name, excel_data in prepared_sheets:
            worksheet = workbook[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.sheet_view.showGridLines = True

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = Font(
                    name="맑은 고딕",
                    size=11,
                    color="FFFFFF",
                    bold=True,
                )
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font

            date_columns = set(find_date_columns(excel_data.columns))
            for column_index, column_name in enumerate(excel_data.columns, start=1):
                letter = worksheet.cell(1, column_index).column_letter
                if column_name == "내역":
                    width = 74
                elif column_name == "주소":
                    width = 38
                elif column_name in {"서비스처리센터", "내역2", "구주소"}:
                    width = 16
                else:
                    width = 13
                worksheet.column_dimensions[letter].width = width
                if column_name in date_columns:
                    for row_index in range(2, worksheet.max_row + 1):
                        worksheet.cell(row_index, column_index).number_format = (
                            "yyyy.mm.dd"
                        )

        workbook.save(temporary_path)
        workbook.close()
        temporary_path.replace(path)
    finally:
        temporary_path.unlink(missing_ok=True)


def save_formatted_excel(
    data: pd.DataFrame,
    path: Path,
    sheet_name: str,
) -> None:
    save_formatted_workbook([(sheet_name, data)], path)


def save_classification_workbook(
    candidates: pd.DataFrame,
    auto_errors: pd.DataFrame,
    path: Path,
) -> None:
    """Store REVIEW and AUTO rows together with an identical column schema."""
    save_formatted_workbook(
        [("후보군", candidates), ("자동오생성", auto_errors)],
        path,
    )


def looks_like_person_name(value: str) -> bool:
    return bool(
        re.fullmatch(r"[가-힣]{2,5}", value)
        and value[0] in KOREAN_SURNAMES
    )


def build_proper_nouns(
    data: pd.DataFrame,
    sap_id_map: dict[str, str],
) -> set[str]:
    proper_nouns = {
        str(name).strip()
        for name in sap_id_map.values()
        if looks_like_person_name(str(name).strip())
    }
    for column in PERSON_NAME_COLUMNS:
        if column not in data.columns:
            continue
        names = data[column].astype("string").str.strip().dropna().unique()
        proper_nouns.update(
            str(name) for name in names if looks_like_person_name(str(name))
        )
    return proper_nouns


def remove_known_proper_nouns(text: str, proper_nouns: set[str]) -> str:
    tokens = text.split()
    kept_tokens: list[str] = []
    for index, token in enumerate(tokens):
        if token in proper_nouns:
            continue

        is_person_token = False
        for suffix in PERSON_SUFFIXES:
            if not token.endswith(suffix) or len(token) <= len(suffix):
                continue
            base = token[: -len(suffix)]
            if base in proper_nouns or (
                looks_like_person_name(base) and base not in NON_PERSON_NAME_BASES
            ):
                is_person_token = True
                break

        previous_token = tokens[index - 1] if index > 0 else ""
        next_token = tokens[index + 1] if index + 1 < len(tokens) else ""
        if (
            looks_like_person_name(token)
            and token not in NON_PERSON_NAME_BASES
            and (
                previous_token in NAME_CONTEXT_TOKENS
                or next_token in NAME_CONTEXT_TOKENS
            )
        ):
            is_person_token = True

        if not is_person_token:
            kept_tokens.append(token)
    return " ".join(kept_tokens)


def normalize_text_value(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    text = text.lower()
    text = re.sub(r"<(?:전화번호|성명|이메일|식별번호)>", " ", text)
    text = re.sub(r"\d+", " ", text)
    text = re.sub(r"[^가-힣a-z]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_details(
    series: pd.Series,
    proper_nouns: set[str],
) -> pd.Series:
    normalized = series.map(normalize_text_value)
    normalized = normalized.map(
        lambda text: remove_known_proper_nouns(text, proper_nouns)
    )
    return normalized.str.replace(r"\s+", " ", regex=True).str.strip()


@dataclass
class PatternLookup:
    singles: dict[str, set[str]]
    phrases: dict[int, dict[tuple[str, ...], set[str]]]
    combo_index: dict[tuple[str, ...], dict[tuple[str, ...], set[str]]]
    combo_lengths: set[int]
    sentences: dict[str, set[str]]


def build_pattern_lookup(documents: Iterable[dict[str, object]]) -> PatternLookup:
    singles: dict[str, set[str]] = defaultdict(set)
    phrases: dict[int, dict[tuple[str, ...], set[str]]] = defaultdict(
        lambda: defaultdict(set)
    )
    combo_index: dict[
        tuple[str, ...], dict[tuple[str, ...], set[str]]
    ] = defaultdict(lambda: defaultdict(set))
    combo_lengths: set[int] = set()
    sentences: dict[str, set[str]] = defaultdict(set)

    for document in documents:
        for owner, patterns_by_type in document.items():
            if not isinstance(patterns_by_type, dict):
                continue
            for pattern_type, keywords in patterns_by_type.items():
                if not isinstance(keywords, list):
                    continue
                for keyword in keywords:
                    if pattern_type == "비연속조합":
                        parts = [part.strip() for part in str(keyword).split("+")]
                        if len(parts) != 2:
                            continue
                        left = tuple(TOKEN_PATTERN.findall(normalize_text_value(parts[0])))
                        right = tuple(TOKEN_PATTERN.findall(normalize_text_value(parts[1])))
                        if not left or not right:
                            continue
                        combo_index[left][right].add(owner)
                        combo_index[right][left].add(owner)
                        combo_lengths.update((len(left), len(right)))
                    elif pattern_type == "반복문장":
                        sentence = normalize_text_value(keyword)
                        if sentence:
                            sentences[sentence].add(owner)
                    else:
                        words = tuple(
                            TOKEN_PATTERN.findall(normalize_text_value(keyword))
                        )
                        if not words:
                            continue
                        if len(words) == 1:
                            singles[words[0]].add(owner)
                        else:
                            phrases[len(words)][words].add(owner)

    return PatternLookup(
        singles=dict(singles),
        phrases={length: dict(values) for length, values in phrases.items()},
        combo_index={
            key: {required: set(owners) for required, owners in values.items()}
            for key, values in combo_index.items()
        },
        combo_lengths=combo_lengths,
        sentences=dict(sentences),
    )


def ngrams(tokens: tuple[str, ...], size: int) -> set[tuple[str, ...]]:
    if size <= 0 or len(tokens) < size:
        return set()
    return {
        tokens[index : index + size]
        for index in range(len(tokens) - size + 1)
    }


def matched_owners(text: str, lookup: PatternLookup) -> set[str]:
    owners: set[str] = set(lookup.sentences.get(text, set()))
    tokens = tuple(TOKEN_PATTERN.findall(text))
    if not tokens:
        return owners

    token_set = set(tokens)
    for token in token_set:
        owners.update(lookup.singles.get(token, set()))

    needed_lengths = set(lookup.phrases) | lookup.combo_lengths
    present_by_length = {
        length: ngrams(tokens, length) for length in needed_lengths
    }

    for length, phrase_owners in lookup.phrases.items():
        for phrase in present_by_length.get(length, set()):
            owners.update(phrase_owners.get(phrase, set()))

    for length in lookup.combo_lengths:
        for present in present_by_length.get(length, set()):
            requirements = lookup.combo_index.get(present)
            if not requirements:
                continue
            for required, required_owners in requirements.items():
                if required in present_by_length.get(len(required), set()):
                    owners.update(required_owners)
    return owners


def load_pattern_documents() -> tuple[dict[str, object], dict[str, object]]:
    keyword_document: dict[str, object] = {}
    conflict_document: dict[str, object] = {}
    if KEYWORD_PATH.is_file():
        with KEYWORD_PATH.open(encoding="utf-8") as file:
            keyword_document = json.load(file)
    if CONFLICT_KEYWORD_PATH.is_file():
        with CONFLICT_KEYWORD_PATH.open(encoding="utf-8") as file:
            conflict_document = json.load(file)
    return keyword_document, conflict_document


def load_priority_pattern_document() -> dict[str, object]:
    """Load owner-only normal rules that must not create cross-owner conflicts."""
    if not PRIORITY_KEYWORD_PATH.is_file():
        return {}
    with PRIORITY_KEYWORD_PATH.open(encoding="utf-8") as file:
        document = json.load(file)
    if not isinstance(document, dict):
        raise ValueError("priority_keyword.json 최상위 값은 객체여야 합니다.")
    return document


def load_auto_normal_pattern_document() -> dict[str, object]:
    """Load only automatically verified normal patterns into the matcher."""
    if not AUTO_NORMAL_PATTERN_PATH.is_file():
        return {}
    with AUTO_NORMAL_PATTERN_PATH.open(encoding="utf-8") as file:
        registry = json.load(file)
    if not isinstance(registry, dict) or not isinstance(
        registry.get("records"), list
    ):
        raise ValueError("auto_normal_pattern.json의 records 배열이 필요합니다.")

    document: dict[str, dict[str, list[str]]] = {}
    for index, record in enumerate(registry["records"]):
        if not isinstance(record, dict):
            raise ValueError(
                f"auto_normal_pattern.json records[{index}]가 객체가 아닙니다."
            )
        if str(record.get("status") or "").strip().casefold() != "active":
            continue
        owner = str(record.get("source_subcategory") or "").strip()
        pattern_type = str(record.get("pattern_type") or "").strip()
        pattern = str(record.get("pattern") or "").strip()
        if not owner or not pattern_type or not pattern:
            raise ValueError(
                "활성 자동 정상 패턴에는 source_subcategory, pattern_type, "
                f"pattern이 필요합니다: records[{index}]"
            )
        values = document.setdefault(owner, {}).setdefault(pattern_type, [])
        if pattern not in values:
            values.append(pattern)
    return document


def load_dominant_normal_pattern_document() -> dict[str, object]:
    """Load strong owner-specific normal phrases for anomaly review only."""
    if not AUTO_NORMAL_PATTERN_PATH.is_file():
        return {}
    with AUTO_NORMAL_PATTERN_PATH.open(encoding="utf-8") as file:
        registry = json.load(file)
    if not isinstance(registry, dict) or not isinstance(
        registry.get("records"), list
    ):
        raise ValueError("auto_normal_pattern.json의 records 배열이 필요합니다.")

    document: dict[str, dict[str, list[str]]] = {}
    for index, record in enumerate(registry["records"]):
        if not isinstance(record, dict):
            raise ValueError(
                f"auto_normal_pattern.json records[{index}]가 객체가 아닙니다."
            )
        status = str(record.get("status") or "").strip().casefold()
        if status not in {"active", "proposed"}:
            continue
        pattern = str(record.get("pattern") or "").strip()
        if len(normalize_text_value(pattern).split()) < 3:
            continue
        if int(record.get("normal_support") or 0) < 10:
            continue
        if int(record.get("other_subcategory_frequency") or 0) != 0:
            continue
        if int(record.get("confirmed_error_pattern_hits") or 0) != 0:
            continue
        owner = str(record.get("source_subcategory") or "").strip()
        pattern_type = str(record.get("pattern_type") or "").strip()
        if not owner or not pattern_type or not pattern:
            continue
        values = document.setdefault(owner, {}).setdefault(pattern_type, [])
        if pattern not in values:
            values.append(pattern)
    return document


def load_scoped_normal_pattern_document() -> dict[str, object]:
    """Load validated subcategory-only patterns that override hard collisions."""
    if not SCOPED_NORMAL_PATTERN_PATH.is_file():
        return {}
    with SCOPED_NORMAL_PATTERN_PATH.open(encoding="utf-8") as file:
        registry = json.load(file)
    if not isinstance(registry, dict) or not isinstance(
        registry.get("records"), list
    ):
        raise ValueError("scoped_normal_pattern.json의 records 배열이 필요합니다.")

    document: dict[str, dict[str, list[str]]] = {}
    for index, record in enumerate(registry["records"]):
        if not isinstance(record, dict):
            raise ValueError(
                f"scoped_normal_pattern.json records[{index}]가 객체가 아닙니다."
            )
        status = str(record.get("status") or "active").strip().casefold()
        if status not in {"active", "enabled", "확정", "정상"}:
            continue
        owner = str(record.get("source_subcategory") or "").strip()
        pattern_type = str(record.get("pattern_type") or "").strip()
        pattern = str(record.get("pattern") or "").strip()
        if not owner or not pattern_type or not pattern:
            raise ValueError(
                "활성 소분류 확정정상 패턴에는 source_subcategory, "
                f"pattern_type, pattern이 필요합니다: records[{index}]"
            )
        values = document.setdefault(owner, {}).setdefault(pattern_type, [])
        if pattern not in values:
            values.append(pattern)
    return document


def without_priority_patterns(
    keyword_document: dict[str, object],
    priority_document: dict[str, object],
) -> dict[str, object]:
    """Remove owner-only normal rules from the cross-owner hard lookup."""
    filtered: dict[str, object] = {}
    for owner, patterns_by_type in keyword_document.items():
        if not isinstance(patterns_by_type, dict):
            continue
        priority_by_type = priority_document.get(owner, {})
        if not isinstance(priority_by_type, dict):
            priority_by_type = {}
        owner_patterns: dict[str, list[object]] = {}
        for pattern_type, keywords in patterns_by_type.items():
            if not isinstance(keywords, list):
                continue
            excluded = {
                str(value)
                for value in priority_by_type.get(pattern_type, [])
            }
            owner_patterns[pattern_type] = [
                value for value in keywords if str(value) not in excluded
            ]
        filtered[owner] = owner_patterns
    return filtered


def precise_conflict_document(
    conflict_document: dict[str, object],
) -> dict[str, object]:
    """Keep concrete conflict phrases/combinations; soften generic singles."""
    result: dict[str, object] = {}
    for owner, patterns_by_type in conflict_document.items():
        if not isinstance(patterns_by_type, dict):
            continue
        result[owner] = {
            pattern_type: keywords
            for pattern_type, keywords in patterns_by_type.items()
            if pattern_type != "단일키워드" and isinstance(keywords, list)
        }
    return result


def load_error_rule_sets(
) -> tuple[
    set[tuple[str, str]],
    set[tuple[str, str]],
    dict[str, dict[int, set[tuple[str, ...]]]],
]:
    """Load exact rules and conservative subcategory-scoped phrase rules."""
    if not ERROR_PATTERN_PATH.is_file():
        if ERROR_PATTERN_PATH == _LEGACY_ERROR_PATTERN_PATH and ERROR_RULES_PATH.is_file():
            from service_order.error.unified_rule_engine import (
                _error_rule_sets,
                _load_json,
            )

            rules = _error_rule_sets(_load_json(ERROR_RULES_PATH))
            return (
                set(rules["active_exact"]),
                set(rules["review_exact"]),
                dict(rules["active_phrases"]),
            )
        return set(), set(), {}
    with ERROR_PATTERN_PATH.open(encoding="utf-8") as file:
        document = json.load(file)
    if not isinstance(document, dict) or not isinstance(document.get("records"), list):
        raise ValueError("error_pattern.json의 records 배열이 필요합니다.")

    active_signatures: set[tuple[str, str]] = set()
    review_signatures: set[tuple[str, str]] = set()
    for index, record in enumerate(document["records"]):
        if not isinstance(record, dict):
            raise ValueError(f"error_pattern.json records[{index}]가 객체가 아닙니다.")
        source = str(record.get("source_subcategory") or "").strip()
        signature = normalize_text_value(record.get("signature"))
        status = str(record.get("status") or "").strip().casefold()
        if not source or not signature or not status:
            raise ValueError(
                "error_pattern.json의 각 record에는 "
                "source_subcategory, signature, status가 필요합니다: "
                f"records[{index}]"
            )
        if status in INACTIVE_ERROR_PATTERN_STATUSES:
            continue
        if status in ACTIVE_ERROR_PATTERN_STATUSES:
            active_signatures.add((source, signature))
            continue
        if status in REVIEW_ERROR_PATTERN_STATUSES:
            review_signatures.add((source, signature))
            continue
        if status not in ACTIVE_ERROR_PATTERN_STATUSES:
            raise ValueError(
                f"error_pattern.json records[{index}]의 status를 해석할 수 없습니다: "
                f"{record.get('status')}"
            )
    overlap = active_signatures.intersection(review_signatures)
    if overlap:
        raise ValueError(
            "error_pattern.json에서 active와 review 서명이 중복됩니다: "
            f"{sorted(overlap)[:3]}"
        )

    active_phrases: dict[str, dict[int, set[tuple[str, ...]]]] = {}
    phrase_records = document.get("phrase_records", [])
    if not isinstance(phrase_records, list):
        raise ValueError("error_pattern.json의 phrase_records는 배열이어야 합니다.")
    for index, record in enumerate(phrase_records):
        if not isinstance(record, dict):
            raise ValueError(
                f"error_pattern.json phrase_records[{index}]가 객체가 아닙니다."
            )
        source = str(record.get("source_subcategory") or "").strip()
        phrase = tuple(normalize_text_value(record.get("phrase")).split())
        status = str(record.get("status") or "").strip().casefold()
        if not source or len(phrase) < 2 or not status:
            raise ValueError(
                "error_pattern.json의 각 phrase_record에는 "
                "source_subcategory, 2어절 이상의 phrase, status가 필요합니다: "
                f"phrase_records[{index}]"
            )
        if status in INACTIVE_ERROR_PATTERN_STATUSES:
            continue
        if status in REVIEW_ERROR_PATTERN_STATUSES:
            # Ambiguous phrases are retained for audit only. Applying a broad
            # ambiguous phrase as REVIEW protection would inflate candidates.
            continue
        if status not in ACTIVE_ERROR_PATTERN_STATUSES:
            raise ValueError(
                "error_pattern.json phrase_records"
                f"[{index}]의 status를 해석할 수 없습니다: {record.get('status')}"
            )
        active_phrases.setdefault(source, {}).setdefault(
            len(phrase), set()
        ).add(phrase)
    return active_signatures, review_signatures, active_phrases


def load_error_signature_sets(
) -> tuple[set[tuple[str, str]], set[tuple[str, str]]]:
    """Backward-compatible exact active/review signature loader."""
    active, review, _ = load_error_rule_sets()
    return active, review


def matches_scoped_error_phrase(
    owner: str,
    signature: str,
    phrase_lookup: dict[str, dict[int, set[tuple[str, ...]]]],
) -> bool:
    normalized_signature = normalize_text_value(signature)
    by_length = phrase_lookup.get(owner, {})
    for phrases in by_length.values():
        for phrase in phrases:
            normalized_phrase = tuple(
                normalize_text_value(token) for token in phrase
            )
            if ordered_distinct_token_match(
                normalized_signature,
                normalized_phrase,
            ):
                return True
    return False


def normalize_error_similarity_template(value: object) -> str:
    """Remove volatile values while retaining the business intent."""
    raw = "" if pd.isna(value) else str(value).lower()
    raw = re.sub(
        r"(?<!\d)01[016789](?:[-.\s]?\d){7,8}(?!\d)",
        " 전화번호값 ",
        raw,
    )
    raw = re.sub(
        r"[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}",
        " 이메일값 ",
        raw,
    )
    text = normalize_text_value(raw)
    if not text:
        return ""
    tokens = text.split()
    normalized: list[str] = []
    amount_tokens = {"원", "만원", "천원", "백만원", "금액"}
    date_tokens = {"일까지", "월까지", "날짜", "일자"}
    period_pattern = re.compile(r"(?:한|두|세|네|몇)?(?:달|개월)(?:분)?$")
    trailing_name_context = (
        "부탁드립니다",
        "요청드립니다",
        "연락처",
        "전화번호",
        "전화번호값",
        "고객명",
        "성명",
        "명의자",
    )
    redact_following_name = False
    for index, token in enumerate(tokens):
        previous = tokens[index - 1] if index else ""
        following = tokens[index + 1] if index + 1 < len(tokens) else ""
        if token in {"전화번호값", "이메일값"}:
            redact_following_name = True
            continue
        probable_context_name = (
            looks_like_person_name(token)
            and token not in NON_PERSON_NAME_BASES
            and (
                redact_following_name
                or
                previous in NAME_CONTEXT_TOKENS
                or following in NAME_CONTEXT_TOKENS
                or previous.endswith(trailing_name_context)
            )
        )
        if probable_context_name:
            redact_following_name = False
            continue
        redact_following_name = False
        if token in amount_tokens:
            normalized.append("금액값")
        elif token in date_tokens:
            normalized.append("날짜값")
        elif period_pattern.fullmatch(token):
            normalized.append("기간값")
        elif token in {"전화번호", "연락처번호"}:
            normalized.append("전화번호값")
        else:
            normalized.append(token)
    return " ".join(normalized)


def load_historical_normal_similarity_signatures(
    path: Path | None = None,
) -> dict[str, set[str]]:
    """Load deduplicated 1~6월 non-error signatures for similarity guards."""
    database = ERROR_LEARNING_BASELINE_PATH if path is None else path
    if not database.is_file():
        return {}
    connection = sqlite3.connect(database)
    try:
        total_rows = connection.execute(
            """
            SELECT source_subcategory, signature, row_count
            FROM signature_counts
            """
        ).fetchall()
        truth_rows = connection.execute(
            """
            SELECT source_subcategory, signature, COUNT(*)
            FROM truth_sources
            GROUP BY source_subcategory, signature
            """
        ).fetchall()
    finally:
        connection.close()
    truth_counts = {
        (str(owner), str(signature)): int(count)
        for owner, signature, count in truth_rows
    }
    result: dict[str, set[str]] = defaultdict(set)
    for owner, signature, total_count in total_rows:
        key = (str(owner), str(signature))
        if int(total_count) - truth_counts.get(key, 0) <= 0:
            continue
        template = normalize_error_similarity_template(signature)
        if template:
            result[str(owner)].add(template)
    return dict(result)


def _similarity_features(
    values: Iterable[str],
) -> list[tuple[str, Counter[str]]]:
    features: list[tuple[str, Counter[str]]] = []
    for value in values:
        text = compact_signature(value)
        if text:
            features.append((text, character_bigram_counter(text)))
    return features


def _top_reference_similarity(
    query: tuple[str, Counter[str]],
    references: list[tuple[str, Counter[str]]],
) -> float:
    if not references:
        return 0.0
    query_text, query_bigrams = query
    preliminary = sorted(
        (
            (counter_cosine(query_bigrams, reference_bigrams), reference_text)
            for reference_text, reference_bigrams in references
        ),
        reverse=True,
    )[:ERROR_SIMILARITY_PREFILTER_LIMIT]
    best = 0.0
    for bigram_score, reference_text in preliminary:
        sequence_score = SequenceMatcher(
            None,
            query_text,
            reference_text,
            autojunk=False,
        ).ratio()
        best = max(best, 0.55 * bigram_score + 0.45 * sequence_score)
    return best


def promote_high_confidence_similar_candidates(
    orders: pd.DataFrame,
    normalized_details: pd.Series,
    candidate_flags: list[bool],
    auto_error_flags: list[bool],
    active_signatures: set[tuple[str, str]],
    review_signatures: set[tuple[str, str]],
) -> int:
    """Promote only candidates far closer to confirmed errors than normals."""
    candidate_indices = []
    for index, is_candidate in enumerate(candidate_flags):
        if not is_candidate:
            continue
        owner = str(orders.iloc[index]["소분류"]).strip()
        signature = str(normalized_details.iloc[index])
        if (owner, signature) in review_signatures:
            continue
        candidate_indices.append(index)
    if not candidate_indices or not active_signatures:
        return 0

    candidate_owners = {
        str(orders.iloc[index]["소분류"]).strip()
        for index in candidate_indices
    }
    positive_by_owner: dict[str, set[str]] = defaultdict(set)
    for owner, signature in active_signatures:
        if owner not in candidate_owners:
            continue
        template = normalize_error_similarity_template(signature)
        if template:
            positive_by_owner[owner].add(template)
    normal_by_owner = load_historical_normal_similarity_signatures()
    positive_features = {
        owner: _similarity_features(values)
        for owner, values in positive_by_owner.items()
    }
    normal_features = {
        owner: _similarity_features(normal_by_owner.get(owner, set()))
        for owner in candidate_owners
    }

    promoted = 0
    for index in candidate_indices:
        owner = str(orders.iloc[index]["소분류"]).strip()
        positives = positive_features.get(owner, [])
        normals = normal_features.get(owner, [])
        if not positives or not normals:
            continue
        template = normalize_error_similarity_template(
            orders.iloc[index]["내역"]
        )
        if (
            informative_token_count(template)
            < ERROR_SIMILARITY_MIN_INFORMATIVE_TOKENS
        ):
            continue
        query_text = compact_signature(template)
        query = (query_text, character_bigram_counter(query_text))
        positive_score = _top_reference_similarity(query, positives)
        if positive_score < ERROR_SIMILARITY_MIN_SCORE:
            continue
        normal_score = _top_reference_similarity(query, normals)
        if normal_score >= ERROR_SIMILARITY_MAX_NORMAL_SCORE:
            continue
        if positive_score - normal_score < ERROR_SIMILARITY_MIN_MARGIN:
            continue
        candidate_flags[index] = False
        auto_error_flags[index] = True
        promoted += 1
    return promoted


def load_confirmed_error_signatures() -> set[tuple[str, str]]:
    """Backward-compatible active confirmed-error loader."""
    return load_error_signature_sets()[0]


def format_classified_orders(orders: pd.DataFrame) -> pd.DataFrame:
    """Apply the common candidate workbook schema to a classified subset."""
    frame = orders.copy()
    if "내역2" not in frame.columns:
        insert_at = frame.columns.get_loc("내역") + 1
        frame.insert(insert_at, "내역2", pd.NA)
    frame = frame.reindex(columns=CANDIDATE_COLUMN_ORDER)
    return frame.sort_values(
        ["오더생성일", "소분류", "오더번호"],
        na_position="last",
    ).reset_index(drop=True)


def _select_candidate_orders_legacy(
    preprocessed: pd.DataFrame,
    sap_id_map: dict[str, str],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, int]]:
    required = {"오더번호", "오더생성일", "소분류", "내역"}
    missing = sorted(required - set(preprocessed.columns))
    if missing:
        raise KeyError(f"후보 판정 필수 열이 없습니다: {missing}")

    subcategories = preprocessed["소분류"].astype("string").str.strip()
    details = preprocessed["내역"].astype("string").str.strip()
    valid_rows = (
        subcategories.notna()
        & subcategories.ne("")
        & details.notna()
        & details.ne("")
    )
    orders = preprocessed.loc[valid_rows].copy().reset_index(drop=True)
    orders["소분류"] = orders["소분류"].astype("string").str.strip()
    orders["내역"] = orders["내역"].astype("string").str.strip()

    proper_nouns = build_proper_nouns(orders, sap_id_map)
    normalized_details = normalize_details(orders["내역"], proper_nouns)
    keyword_document, conflict_document = load_pattern_documents()
    priority_document = load_priority_pattern_document()
    auto_normal_document = load_auto_normal_pattern_document()
    dominant_normal_document = load_dominant_normal_pattern_document()
    scoped_normal_document = load_scoped_normal_pattern_document()
    keyword_lookup = build_pattern_lookup(
        [
            keyword_document,
            priority_document,
            auto_normal_document,
            scoped_normal_document,
        ]
    )
    scoped_normal_lookup = build_pattern_lookup(
        [auto_normal_document, scoped_normal_document]
    )
    collision_lookup = build_pattern_lookup(
        [
            keyword_document,
            priority_document,
            auto_normal_document,
            scoped_normal_document,
            conflict_document,
        ]
    )
    hard_collision_lookup = build_pattern_lookup(
        [
            without_priority_patterns(keyword_document, priority_document),
            precise_conflict_document(conflict_document),
        ]
    )
    dominant_normal_lookup = build_pattern_lookup([dominant_normal_document])
    (
        confirmed_error_signatures,
        review_error_signatures,
        confirmed_error_phrases,
    ) = load_error_rule_sets()

    candidate_flags: list[bool] = []
    auto_error_flags: list[bool] = []
    own_match_count = 0
    other_match_count = 0
    hard_other_match_count = 0
    softened_collision_count = 0
    scoped_normal_match_count = 0
    scoped_collision_override_count = 0
    protected_review_count = 0
    dominant_normal_anomaly_count = 0
    exact_auto_error_count = 0
    phrase_auto_error_count = 0
    normal_count = 0
    for subcategory, detail in zip(
        orders["소분류"], normalized_details
    ):
        owner = "" if pd.isna(subcategory) else str(subcategory)
        exact_auto_error = (owner, detail) in confirmed_error_signatures
        phrase_auto_error = (
            not exact_auto_error
            and matches_scoped_error_phrase(
                owner,
                detail,
                confirmed_error_phrases,
            )
        )
        auto_error = exact_auto_error or phrase_auto_error
        protected_review = (owner, detail) in review_error_signatures
        dominant_normal_owners = matched_owners(detail, dominant_normal_lookup)
        dominant_normal_anomaly = any(
            pattern_owner != owner
            for pattern_owner in dominant_normal_owners
        )
        keyword_owners = matched_owners(detail, keyword_lookup)
        scoped_normal_owners = matched_owners(detail, scoped_normal_lookup)
        collision_owners = matched_owners(detail, collision_lookup)
        hard_collision_owners = matched_owners(detail, hard_collision_lookup)
        own_match = owner in keyword_owners
        scoped_normal_match = owner in scoped_normal_owners
        other_match = any(pattern_owner != owner for pattern_owner in collision_owners)
        hard_other_match = any(
            pattern_owner != owner for pattern_owner in hard_collision_owners
        )
        normal = (
            not auto_error
            and not protected_review
            and not dominant_normal_anomaly
            and (
                scoped_normal_match
                or (own_match and not hard_other_match)
            )
        )
        candidate_flags.append(not auto_error and not normal)
        auto_error_flags.append(auto_error)
        own_match_count += int(own_match)
        other_match_count += int(other_match)
        hard_other_match_count += int(hard_other_match)
        softened_collision_count += int(normal and other_match)
        scoped_normal_match_count += int(scoped_normal_match)
        scoped_collision_override_count += int(
            normal and scoped_normal_match and hard_other_match
        )
        protected_review_count += int(protected_review)
        dominant_normal_anomaly_count += int(dominant_normal_anomaly)
        exact_auto_error_count += int(exact_auto_error)
        phrase_auto_error_count += int(phrase_auto_error)
        normal_count += int(normal)

    similarity_auto_error_count = promote_high_confidence_similar_candidates(
        orders,
        normalized_details,
        candidate_flags,
        auto_error_flags,
        confirmed_error_signatures,
        review_error_signatures,
    )
    candidates = format_classified_orders(orders.loc[candidate_flags])
    auto_errors = format_classified_orders(orders.loc[auto_error_flags])

    return candidates, auto_errors, {
        "분석대상행수": len(orders),
        "빈내역제외행수": int((~valid_rows).sum()),
        "자기패턴일치행수": own_match_count,
        "타분류충돌행수": other_match_count,
        "구체충돌행수": hard_other_match_count,
        "일반단일충돌약화행수": softened_collision_count,
        "소분류확정정상일치행수": scoped_normal_match_count,
        "소분류확정정상충돌우선행수": scoped_collision_override_count,
        "확정검토보호행수": protected_review_count,
        "소분류지배문구이상행수": dominant_normal_anomaly_count,
        "정확문장자동오생성행수": exact_auto_error_count,
        "문구조합자동오생성행수": phrase_auto_error_count,
        "유사문장자동오생성행수": similarity_auto_error_count,
        "정상제외행수": normal_count,
        "자동오생성행수": len(auto_errors),
        "검토후보행수": len(candidates),
        "후보행수": len(candidates),
    }


def _legacy_rule_fixture_active() -> bool:
    """Allow historical tests to patch temporary legacy registries.

    The production paths no longer depend on those files.  A non-default path
    means an explicit test/migration fixture requested the old classifier.
    """
    return any(
        current != default
        for current, default in (
            (KEYWORD_PATH, _LEGACY_KEYWORD_PATH),
            (CONFLICT_KEYWORD_PATH, _LEGACY_CONFLICT_KEYWORD_PATH),
            (PRIORITY_KEYWORD_PATH, _LEGACY_PRIORITY_KEYWORD_PATH),
            (ERROR_PATTERN_PATH, _LEGACY_ERROR_PATTERN_PATH),
            (AUTO_NORMAL_PATTERN_PATH, _LEGACY_AUTO_NORMAL_PATTERN_PATH),
            (SCOPED_NORMAL_PATTERN_PATH, _LEGACY_SCOPED_NORMAL_PATTERN_PATH),
        )
    )


def select_candidate_orders(
    preprocessed: pd.DataFrame,
    sap_id_map: dict[str, str],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, object]]:
    """Classify with the consolidated runtime rules.

    The import is intentionally local because unified_rule_engine reuses the
    normalization and formatting helpers defined in this module.
    """
    if _legacy_rule_fixture_active():
        return _select_candidate_orders_legacy(preprocessed, sap_id_map)
    from service_order.error.unified_rule_engine import (
        select_candidate_orders_unified,
    )

    return select_candidate_orders_unified(
        preprocessed,
        enable_context_policies=True,
    )


def _order_number_keys(frame: pd.DataFrame) -> set[str]:
    if "오더번호" not in frame.columns:
        return set()
    return set(frame["오더번호"].astype("string").fillna("").tolist()) - {""}


def _contains_probable_person_name(
    pattern: tuple[str, ...],
    proper_nouns: set[str],
) -> bool:
    """Reject person-bearing phrases from automatic normal-rule promotion."""
    name_context = NAME_CONTEXT_TOKENS | {
        "배우자", "아들", "딸", "본인", "명의자", "집주인", "고객", "님",
        "매니저님", "기사님", "과장님", "대리님", "사원님", "사무실오더",
        "부적합", "요청자", "고객명", "성명", "명의",
    }
    for index, token in enumerate(pattern):
        if token in proper_nouns:
            return True
        for suffix in PERSON_SUFFIXES:
            if not token.endswith(suffix) or len(token) <= len(suffix):
                continue
            base = token[: -len(suffix)]
            if looks_like_person_name(base) and base not in NON_PERSON_NAME_BASES:
                return True
        previous = pattern[index - 1] if index else ""
        following = pattern[index + 1] if index + 1 < len(pattern) else ""
        if (
            looks_like_person_name(token)
            and token not in NON_PERSON_NAME_BASES
            and (previous in name_context or following in name_context)
        ):
            return True
    return False


def _active_registry_record_count(path: Path) -> int:
    if not path.is_file():
        return 0
    document = json.loads(path.read_text(encoding="utf-8"))
    records = document.get("records", []) if isinstance(document, dict) else []
    if not isinstance(records, list):
        return 0
    return sum(
        isinstance(record, dict)
        and str(record.get("status") or "active").strip().casefold()
        in {"active", "enabled", "확정", "정상"}
        for record in records
    )


def _active_unified_override_rule_count(path: Path = NORMAL_RULES_PATH) -> int:
    if not path.is_file():
        return 0
    document = json.loads(path.read_text(encoding="utf-8"))
    rules = document.get("rules", []) if isinstance(document, dict) else []
    if not isinstance(rules, list):
        return 0
    return sum(
        isinstance(record, dict)
        and isinstance(record.get("behavior"), dict)
        and bool(record["behavior"].get("override_collision"))
        for record in rules
    )


def update_auto_normal_pattern_registry(
    preprocessed: pd.DataFrame,
    candidates: pd.DataFrame,
    auto_errors: pd.DataFrame,
    sap_id_map: dict[str, str],
) -> dict[str, object]:
    """Accumulate aggregate normal evidence without retaining uploaded rows.

    Evidence snapshots are date-range aggregates. Overlapping uploads replace
    prior snapshots, so re-uploading the same month cannot inflate support.
    A two-token phrase may now be promoted when it has stronger, diverse normal
    evidence and affects only a small review group. Three-or-more-token phrases
    need less support because their meaning is more specific. A repeatedly
    observed owner-only phrase can still be promoted through the stricter
    high-volume path. Every path requires zero foreign-category frequency and
    zero approved-error evidence.
    """
    required = {"오더번호", "오더생성일", "소분류", "내역"}
    if missing := sorted(required - set(preprocessed.columns)):
        raise KeyError(f"자동 정상 패턴 수집 필수 열이 없습니다: {missing}")

    frame = preprocessed.loc[
        preprocessed["소분류"].astype("string").str.strip().fillna("").ne("")
        & preprocessed["내역"].astype("string").str.strip().fillna("").ne("")
    ].copy()
    if frame.empty:
        return {
            "active_changed": False,
            "active_total": 0,
            "proposed_total": 0,
            "new_active": 0,
            "demoted": 0,
            "current_candidate_impact": 0,
            "evidence_snapshot_total": 0,
            "evidence_pattern_total": 0,
        }

    proper_nouns = build_proper_nouns(frame, sap_id_map)
    frame["normal_text"] = normalize_details(frame["내역"], proper_nouns)
    frame["order_key"] = frame["오더번호"].astype("string").fillna("")
    frame["pattern_owner"] = frame["소분류"].astype("string").str.strip().fillna("")
    frame["order_date_key"] = pd.to_datetime(
        frame["오더생성일"], errors="coerce"
    ).dt.date.astype("string")
    if "생성인" in frame.columns:
        frame["person_key"] = frame["생성인"].astype("string").str.strip().fillna("")
    else:
        frame["person_key"] = ""

    candidate_keys = _order_number_keys(candidates)
    auto_error_keys = _order_number_keys(auto_errors)
    frame["classification"] = "normal"
    frame.loc[frame["order_key"].isin(candidate_keys), "classification"] = "candidate"
    frame.loc[frame["order_key"].isin(auto_error_keys), "classification"] = "error"

    total_by_pattern: dict[tuple[str, ...], dict[str, int]] = defaultdict(
        lambda: defaultdict(int)
    )
    normal_count: defaultdict[tuple[str, tuple[str, ...]], int] = defaultdict(int)
    candidate_rows: dict[tuple[str, tuple[str, ...]], set[str]] = defaultdict(set)
    normal_dates: dict[tuple[str, tuple[str, ...]], set[str]] = defaultdict(set)
    normal_people: dict[tuple[str, tuple[str, ...]], set[str]] = defaultdict(set)

    active_errors, review_errors = load_error_signature_sets()
    error_tokens_by_owner: dict[str, list[tuple[str, ...]]] = defaultdict(list)
    for error_owner, signature in active_errors | review_errors:
        error_tokens_by_owner[error_owner].append(
            tuple(TOKEN_PATTERN.findall(signature))
        )
    error_hit_cache: dict[tuple[str, tuple[str, ...]], int] = {}

    def error_pattern_hits(owner: str, pattern: tuple[str, ...]) -> int:
        key = (owner, pattern)
        if key not in error_hit_cache:
            size = len(pattern)
            error_hit_cache[key] = sum(
                pattern in ngrams(tokens, size)
                for tokens in error_tokens_by_owner.get(owner, [])
            )
        return error_hit_cache[key]

    for row in frame.itertuples(index=False):
        owner = str(getattr(row, "pattern_owner"))
        tokens = tuple(TOKEN_PATTERN.findall(str(getattr(row, "normal_text"))))
        classification = str(getattr(row, "classification"))
        order_key = str(getattr(row, "order_key"))
        row_patterns = {
            pattern
            for size in range(2, min(4, len(tokens)) + 1)
            for pattern in ngrams(tokens, size)
            if not _contains_probable_person_name(pattern, proper_nouns)
        }
        for pattern in row_patterns:
            key = (owner, pattern)
            total_by_pattern[pattern][owner] += 1
            if classification == "normal":
                normal_count[key] += 1
                normal_dates[key].add(str(getattr(row, "order_date_key")))
                person = str(getattr(row, "person_key"))
                if person:
                    normal_people[key].add(person)
            elif classification == "candidate":
                candidate_rows[key].add(order_key)

    previous_document: dict[str, object]
    if _legacy_rule_fixture_active():
        previous_document = {
            "version": 2,
            "records": [],
            "evidence_snapshots": [],
        }
        if AUTO_NORMAL_PATTERN_PATH.is_file():
            previous_document = json.loads(
                AUTO_NORMAL_PATTERN_PATH.read_text(encoding="utf-8")
            )
    else:
        from service_order.error.unified_rule_registry import (
            extract_auto_normal_registry,
        )

        previous_document = extract_auto_normal_registry(NORMAL_RULES_PATH)
    previous_records = previous_document.get("records", [])
    if not isinstance(previous_records, list):
        raise ValueError("auto_normal_pattern.json의 records 배열이 필요합니다.")

    records_by_key: dict[tuple[str, tuple[str, ...]], dict[str, object]] = {}
    previous_active: set[tuple[str, tuple[str, ...]]] = set()
    for record in previous_records:
        if not isinstance(record, dict):
            continue
        owner = str(record.get("source_subcategory") or "").strip()
        pattern = tuple(
            TOKEN_PATTERN.findall(normalize_text_value(record.get("pattern")))
        )
        if not owner or len(pattern) < 2:
            continue
        key = (owner, pattern)
        records_by_key[key] = dict(record)
        if str(record.get("status") or "").casefold() == "active":
            previous_active.add(key)

    dates = pd.to_datetime(frame["오더생성일"], errors="coerce").dropna()
    if dates.empty:
        raise ValueError("자동 정상 패턴의 증거 기간을 계산할 수 없습니다.")
    snapshot_start = dates.min().date().isoformat()
    snapshot_end = dates.max().date().isoformat()
    snapshot_updated_at = datetime.now().astimezone().isoformat()

    snapshot_records: list[dict[str, object]] = []
    for key, support in normal_count.items():
        if support < 2:
            continue
        owner, pattern = key
        other_count = sum(
            count
            for pattern_owner, count in total_by_pattern[pattern].items()
            if pattern_owner != owner
        )
        snapshot_records.append(
            {
                "source_subcategory": owner,
                "pattern": " ".join(pattern),
                "normal_support": int(support),
                "candidate_coverage": len(candidate_rows.get(key, set())),
                "other_subcategory_frequency": int(other_count),
                "distinct_normal_dates": len(normal_dates.get(key, set()) - {"<NA>"}),
                "distinct_normal_people": len(normal_people.get(key, set())),
            }
        )

    previous_snapshots = previous_document.get("evidence_snapshots", [])
    if not isinstance(previous_snapshots, list):
        raise ValueError("auto_normal_pattern.json의 evidence_snapshots 배열이 필요합니다.")

    def overlaps(snapshot: object) -> bool:
        if not isinstance(snapshot, dict):
            return False
        start = str(snapshot.get("start") or "")
        end = str(snapshot.get("end") or "")
        return bool(start and end and start <= snapshot_end and snapshot_start <= end)

    evidence_snapshots = [
        snapshot
        for snapshot in previous_snapshots
        if isinstance(snapshot, dict) and not overlaps(snapshot)
    ]
    evidence_snapshots.append(
        {
            "start": snapshot_start,
            "end": snapshot_end,
            "updated_at": snapshot_updated_at,
            "patterns": snapshot_records,
        }
    )
    evidence_snapshots = sorted(
        evidence_snapshots,
        key=lambda snapshot: (str(snapshot.get("start") or ""), str(snapshot.get("end") or "")),
    )[-36:]

    cumulative: dict[tuple[str, tuple[str, ...]], dict[str, int]] = defaultdict(
        lambda: defaultdict(int)
    )
    snapshot_counts: defaultdict[tuple[str, tuple[str, ...]], int] = defaultdict(int)
    for snapshot in evidence_snapshots:
        patterns = snapshot.get("patterns", [])
        if not isinstance(patterns, list):
            continue
        for evidence in patterns:
            if not isinstance(evidence, dict):
                continue
            owner = str(evidence.get("source_subcategory") or "").strip()
            pattern = tuple(
                TOKEN_PATTERN.findall(normalize_text_value(evidence.get("pattern")))
            )
            if not owner or len(pattern) < 2:
                continue
            key = (owner, pattern)
            values = cumulative[key]
            values["normal_support"] += int(evidence.get("normal_support") or 0)
            values["candidate_coverage"] += int(evidence.get("candidate_coverage") or 0)
            values["other_subcategory_frequency"] += int(
                evidence.get("other_subcategory_frequency") or 0
            )
            values["distinct_normal_dates"] += int(
                evidence.get("distinct_normal_dates") or 0
            )
            values["distinct_normal_people"] = max(
                values["distinct_normal_people"],
                int(evidence.get("distinct_normal_people") or 0),
            )
            snapshot_counts[key] += 1

    proposals = sorted(
        (
            key
            for key, values in cumulative.items()
            if values["normal_support"] >= 5
        ),
        key=lambda key: (
            -len(key[1]),
            -cumulative[key]["normal_support"],
            -cumulative[key]["distinct_normal_dates"],
            " ".join(key[1]),
        ),
    )

    relieved_candidate_rows: set[str] = set()
    repeated_candidate_promotions = 0
    repeated_promoted_patterns: defaultdict[str, list[tuple[str, ...]]] = defaultdict(
        list
    )
    selected_new_active: set[tuple[str, tuple[str, ...]]] = set()

    def is_nested_phrase(
        phrase: tuple[str, ...],
        longer_phrase: tuple[str, ...],
    ) -> bool:
        return any(
            longer_phrase[index : index + len(phrase)] == phrase
            for index in range(len(longer_phrase) - len(phrase) + 1)
        )

    for key in proposals:
        owner, pattern = key
        if key in previous_active or len(pattern) < 2:
            continue
        rows = candidate_rows.get(key, set())
        values = cumulative[key]
        common_safe_conditions = (
            values["other_subcategory_frequency"] == 0
            and error_pattern_hits(owner, pattern) == 0
        )
        if len(pattern) == 2:
            normal_support_threshold = 8
            normal_date_threshold = 3
            normal_people_threshold = 2
        else:
            normal_support_threshold = 5
            normal_date_threshold = 2
            normal_people_threshold = 2
        evidence_backed_candidate_promotion = (
            values["normal_support"] >= normal_support_threshold
            and values["distinct_normal_dates"] >= normal_date_threshold
            and values["distinct_normal_people"] >= normal_people_threshold
            and bool(rows)
            and len(rows) <= 3
        )
        repeated_candidate_promotion = (
            values["normal_support"] >= 20
            and values["distinct_normal_dates"] >= 3
            and values["distinct_normal_people"] >= 2
            and values["candidate_coverage"] >= 5
        )
        if not common_safe_conditions or not (
            evidence_backed_candidate_promotion or repeated_candidate_promotion
        ):
            continue
        if any(
            is_nested_phrase(pattern, existing)
            for existing in repeated_promoted_patterns[owner]
        ):
            continue
        selected_new_active.add(key)
        relieved_candidate_rows.update(rows)
        if not evidence_backed_candidate_promotion:
            repeated_candidate_promotions += 1
        repeated_promoted_patterns[owner].append(pattern)

    all_keys = set(proposals) | previous_active
    current_active: set[tuple[str, tuple[str, ...]]] = set()
    output_records: list[dict[str, object]] = []
    for key in sorted(all_keys, key=lambda value: (value[0], value[1])):
        owner, pattern = key
        previous = records_by_key.get(key, {})
        values = cumulative.get(key, {})
        support = int(values.get("normal_support", 0))
        rows = candidate_rows.get(key, set())
        other_count = int(values.get("other_subcategory_frequency", 0))
        error_hits = error_pattern_hits(owner, pattern)
        was_active = key in previous_active
        if error_hits:
            status = "blocked_error"
        elif was_active and other_count == 0:
            status = "active"
        elif key in selected_new_active:
            status = "active"
        else:
            status = "proposed"
        if status == "active":
            current_active.add(key)
        output_records.append(
            {
                "source_subcategory": owner,
                "pattern_type": f"{len(pattern)}어절문구",
                "pattern": " ".join(pattern),
                "status": status,
                "normal_support": int(support),
                "candidate_coverage": int(values.get("candidate_coverage", 0)),
                "current_candidate_coverage": len(rows),
                "other_subcategory_frequency": int(other_count),
                "distinct_normal_dates": int(values.get("distinct_normal_dates", 0)),
                "distinct_normal_people": int(values.get("distinct_normal_people", 0)),
                "evidence_snapshot_count": int(snapshot_counts.get(key, 0)),
                "confirmed_error_pattern_hits": int(error_hits),
                "first_seen_at": previous.get(
                    "first_seen_at", datetime.now().astimezone().isoformat()
                ),
                "last_evaluated_at": datetime.now().astimezone().isoformat(),
            }
        )

    document = {
        "version": 2,
        "policy": {
            "evidence_backed_candidate_relief": {
                "two_token": {
                    "active_min_normal_support": 8,
                    "active_min_distinct_dates": 3,
                    "active_min_distinct_people": 2,
                },
                "three_or_more_token": {
                    "active_min_normal_support": 5,
                    "active_min_distinct_dates": 2,
                    "active_min_distinct_people": 2,
                },
                "maximum_current_candidate_coverage": 3,
                "same_subcategory_only": True,
                "other_subcategory_frequency": 0,
                "confirmed_error_pattern_hits": 0,
            },
            "repeated_candidate_relief": {
                "active_min_normal_support": 20,
                "active_min_distinct_dates": 3,
                "active_min_distinct_people": 2,
                "active_min_cumulative_candidate_coverage": 5,
                "same_subcategory_only": True,
                "confirmed_error_pattern_hits": 0,
            },
            "required_other_subcategory_frequency": 0,
            "required_confirmed_error_pattern_hits": 0,
            "overlapping_period_policy": "replace",
            "maximum_evidence_snapshots": 36,
        },
        "evidence_snapshots": evidence_snapshots,
        "records": output_records,
    }
    unified_merge_summary: dict[str, object] = {}
    if _legacy_rule_fixture_active():
        AUTO_NORMAL_PATTERN_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = AUTO_NORMAL_PATTERN_PATH.with_name(
            f".{AUTO_NORMAL_PATTERN_PATH.stem}_{uuid.uuid4().hex[:8]}_writing.json"
        )
        try:
            temporary_path.write_text(
                json.dumps(document, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            temporary_path.replace(AUTO_NORMAL_PATTERN_PATH)
        finally:
            temporary_path.unlink(missing_ok=True)
    else:
        from service_order.error.unified_rule_registry import (
            merge_auto_normal_registry,
        )

        unified_merge_summary = merge_auto_normal_registry(
            document,
            path=NORMAL_RULES_PATH,
            training_cutoff=snapshot_end,
        )

    return {
        "active_changed": current_active != previous_active,
        "active_total": len(current_active),
        "proposed_total": sum(
            record["status"] == "proposed" for record in output_records
        ),
        "new_active": len(current_active - previous_active),
        "demoted": len(previous_active - current_active),
        "current_candidate_impact": len(relieved_candidate_rows),
        "repeated_candidate_promotions": repeated_candidate_promotions,
        "evidence_snapshot_total": len(evidence_snapshots),
        "evidence_pattern_total": sum(
            len(snapshot.get("patterns", []))
            for snapshot in evidence_snapshots
            if isinstance(snapshot, dict)
        ),
        "unified_runtime_revision": int(
            unified_merge_summary.get("runtime_revision") or 0
        ),
    }


def candidate_period(data: pd.DataFrame) -> tuple[str, str, str]:
    dates = pd.to_datetime(data["오더생성일"], errors="coerce").dropna()
    if dates.empty:
        raise ValueError("오더생성일이 없어 결과 기간을 만들 수 없습니다.")
    start_date = dates.min()
    end_date = dates.max()
    if start_date.month == end_date.month:
        label = f"{start_date.month}_{start_date.day}-{end_date.day}"
    else:
        label = (
            f"{start_date.month}_{start_date.day}-"
            f"{end_date.month}_{end_date.day}"
        )
    return label, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")


def run_analysis(
    source_path: Path,
    output_dir: Path,
    original_stem: str,
    *,
    progress_callback: Callable[[int, str], None] | None = None,
) -> dict[str, object]:
    def report(percent: int, message: str) -> None:
        if progress_callback is not None:
            progress_callback(max(0, min(100, int(percent))), message)

    report(20, "분석 기준과 키워드 규칙을 불러오고 있습니다.")
    sap_id_map, except_keys = load_configuration()
    report(25, "업로드한 Excel 데이터를 읽고 있습니다.")
    raw, source_summary = read_order_excel(source_path)
    report(36, "원본 데이터 전처리를 진행하고 있습니다.")
    preprocessed, dashboard_totals, preprocess_summary = preprocess_orders(
        raw,
        sap_id_map,
        except_keys,
        return_dashboard_totals=True,
    )
    report(52, "전처리 결과의 열과 데이터 형식을 확인하고 있습니다.")
    validate_preprocessed(preprocessed)
    validate_preprocessed(dashboard_totals)
    preprocess_summary = {**source_summary, **preprocess_summary}
    report(60, "정상 규칙과 오생성 패턴으로 후보를 분류하고 있습니다.")
    candidates, auto_errors, candidate_summary = select_candidate_orders(
        preprocessed,
        sap_id_map,
    )
    report(69, "반복 정상 패턴을 갱신하고 후보를 정리하고 있습니다.")
    auto_normal_summary = update_auto_normal_pattern_registry(
        preprocessed,
        candidates,
        auto_errors,
        sap_id_map,
    )
    if auto_normal_summary["active_changed"]:
        report(74, "새로 활성화된 정상 패턴을 반영해 다시 분류하고 있습니다.")
        candidates, auto_errors, candidate_summary = select_candidate_orders(
            preprocessed,
            sap_id_map,
        )
    candidate_summary = {
        **candidate_summary,
        "자동정상활성패턴수": int(auto_normal_summary["active_total"]),
        "자동정상추천패턴수": int(auto_normal_summary["proposed_total"]),
        "자동정상신규활성수": int(auto_normal_summary["new_active"]),
        "자동정상비활성전환수": int(auto_normal_summary["demoted"]),
        "자동정상당회후보감소추정": int(
            auto_normal_summary["current_candidate_impact"]
        ),
        "자동정상누적기간수": int(
            auto_normal_summary["evidence_snapshot_total"]
        ),
        "자동정상누적증거패턴수": int(
            auto_normal_summary["evidence_pattern_total"]
        ),
        "소분류확정정상패턴수": _active_unified_override_rule_count(),
    }
    # Classification and normal-pattern learning use the in-memory source
    # wording. Only after those decisions are complete do exposed/stored
    # frames receive irreversible personal-data masking.
    preprocessed, privacy_summary = mask_personal_data_frame(preprocessed)
    dashboard_totals, _ = mask_personal_data_frame(dashboard_totals)
    candidates, _ = mask_personal_data_frame(candidates)
    auto_errors, _ = mask_personal_data_frame(auto_errors)
    preprocess_summary = {**preprocess_summary, **privacy_summary}
    report(80, "분석 기간과 결과 파일 구성을 확정하고 있습니다.")
    period, start_date, end_date = candidate_period(preprocessed)

    output_dir.mkdir(parents=True, exist_ok=True)
    preprocessed_path = output_dir / f"{original_stem}_total_data.xlsx"
    candidate_path = output_dir / f"{period}_candidate.xlsx"
    report(84, "전체 데이터 Excel을 만들고 있습니다.")
    save_formatted_excel(preprocessed, preprocessed_path, "전체데이터")
    report(89, "오생성과 후보 Excel을 만들고 있습니다.")
    save_classification_workbook(candidates, auto_errors, candidate_path)
    report(92, "분석 결과를 서버에 반영할 준비를 하고 있습니다.")

    return {
        "preprocessed_path": preprocessed_path,
        "candidate_path": candidate_path,
        # The web server keeps these frames only for the current analysis job.
        # They are never written to the dashboard database.
        "preprocessed": preprocessed,
        "dashboard_totals": dashboard_totals,
        "candidates": candidates,
        "auto_errors": auto_errors,
        "start_date": start_date,
        "end_date": end_date,
        "preprocess_summary": preprocess_summary,
        "candidate_summary": candidate_summary,
    }
