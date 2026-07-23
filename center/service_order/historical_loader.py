"""Read historical service-order data and return aggregate dashboard metrics.

This module intentionally has no persistence code. It reads one monthly source
workbook at a time, reduces it to aggregate counts, and discards the detailed
rows before the next month is opened. Legacy pickle inputs remain supported for
tests and migrations, but the dashboard uses the original monthly workbooks.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from numbers import Integral, Real
from pathlib import Path
from typing import Iterable, Mapping
import json
import re

import pandas as pd
from openpyxl import load_workbook

from .error.analysis_pipeline import (
    BUSINESS_BY_CREATOR_PREFIX,
    CREATOR_BUSINESS_EXCEPTIONS,
)


ORDER_NUMBER = "오더번호"
ORDER_DATE = "오더생성일"
STATUS = "상태"
PERSON = "생성인"
DEPARTMENT = "생성부서"
BUSINESS = "사업부"
SUBCATEGORY = "소분류"
SERVICE_CENTER = "서비스처리센터"
BUSINESS_DEPARTMENT = "사업부"
BUSINESS_VALUES = frozenset({"중부", "북부", "남부", "동부", "서부"})
BUSINESS_BY_CENTER = {
    "H071": "중부",
    "H072": "북부",
    "H073": "남부",
    "H074": "동부",
    "H075": "서부",
}
CANCELLED_ORDER_STATUS = "오더취소"
UNKNOWN_VALUE = "(미확인)"
ORDER_CREATOR = "오더생성자"
CUSTOMER_SERVICE_CENTER = "고객서비스처리센터"
TOTAL_HEADER_SIGNATURE = frozenset(
    {ORDER_NUMBER, ORDER_DATE, STATUS, SUBCATEGORY, ORDER_CREATOR}
)
MAX_TOTAL_HEADER_ROWS = 30

TRUTH_COLUMNS = (
    ORDER_NUMBER,
    ORDER_DATE,
    STATUS,
    PERSON,
    DEPARTMENT,
    BUSINESS,
    SUBCATEGORY,
)
TOTAL_COLUMNS = TRUTH_COLUMNS
TOTAL_SOURCE_COLUMNS = frozenset(
    {
        *TOTAL_COLUMNS,
        ORDER_CREATOR,
        SERVICE_CENTER,
        CUSTOMER_SERVICE_CENTER,
    }
)
AGGREGATE_COLUMNS = (
    "metric_date",
    "person",
    "business",
    "subcategory",
    "total_count",
    "error_count",
)


class HistoricalDataError(RuntimeError):
    """Base exception for historical data that cannot be loaded safely."""


class HistoricalInputError(HistoricalDataError):
    """Raised when an input file, sheet, or required column is missing."""


class HistoricalValidationError(HistoricalDataError):
    """Raised when historical truth and total data do not reconcile."""


@dataclass(frozen=True, slots=True)
class MonthLoadSummary:
    """Counts produced while reducing one monthly total-data pickle."""

    month: int
    source_rows: int
    total_count: int
    excluded_department_rows: int
    excluded_business_rows: int
    excluded_subcategory_rows: int
    excluded_cancelled_rows: int
    truth_source_rows: int
    truth_count: int
    truth_excluded_department_rows: int
    truth_excluded_business_rows: int
    truth_excluded_subcategory_rows: int
    truth_excluded_cancelled_rows: int
    matched_truth_count: int
    aggregate_rows: int


@dataclass(frozen=True, slots=True)
class HistoricalLoadSummary:
    """Validation and row-count summary for a historical load."""

    months: tuple[int, ...]
    source_total_rows: int
    total_count: int
    excluded_department_rows: int
    excluded_business_rows: int
    excluded_subcategory_rows: int
    excluded_cancelled_rows: int
    truth_source_rows: int
    truth_count: int
    truth_excluded_department_rows: int
    truth_excluded_business_rows: int
    truth_excluded_subcategory_rows: int
    truth_excluded_cancelled_rows: int
    truth_missing_order_rows: int
    truth_duplicate_order_count: int
    truth_unmatched_count: int
    truth_dimension_mismatch_count: int
    aggregate_rows: int
    month_summaries: tuple[MonthLoadSummary, ...]

    def as_dict(self) -> dict[str, object]:
        """Return a JSON-serializable representation of this summary."""

        return asdict(self)


@dataclass(slots=True)
class HistoricalLoadResult:
    """Aggregate-only historical result returned to the dashboard layer."""

    aggregates: pd.DataFrame
    summary: HistoricalLoadSummary


@dataclass(frozen=True, slots=True)
class _TruthRecord:
    order_number: str
    metric_date: str
    person: str
    business: str
    subcategory: str
    month: int
    sheet: str
    row_number: int


@dataclass(frozen=True, slots=True)
class _TruthSheetSummary:
    source_rows: int = 0
    truth_count: int = 0
    excluded_department_rows: int = 0
    excluded_business_rows: int = 0
    excluded_subcategory_rows: int = 0
    excluded_cancelled_rows: int = 0
    missing_order_rows: int = 0


def _clean_scalar(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _dimension_scalar(value: object) -> str:
    value = _clean_scalar(value)
    return value or UNKNOWN_VALUE


def _subcategory_key(value: object) -> str:
    return re.sub(r"\s+", "", _clean_scalar(value))


def _normalize_order_number(value: object) -> str | None:
    """Normalize Excel/pandas numeric order IDs without losing string IDs."""

    if value is None or pd.isna(value):
        return None
    if isinstance(value, Integral) and not isinstance(value, bool):
        return str(int(value))
    if isinstance(value, Real) and not isinstance(value, bool):
        number = float(value)
        if not pd.notna(number):
            return None
        if number.is_integer():
            return str(int(number))

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if re.fullmatch(r"[+-]?\d+\.0+", text):
        return text.split(".", 1)[0].lstrip("+")
    return text


def _normalize_order_series(series: pd.Series) -> pd.Series:
    return series.map(_normalize_order_number).astype("string")


def _normalize_date_scalar(value: object) -> str | None:
    if value is None or pd.isna(value) or _clean_scalar(value) == "":
        return None
    converted = pd.to_datetime(value, errors="coerce")
    if pd.isna(converted):
        return None
    return converted.date().isoformat()


def _load_exclusion_keys(path: Path) -> set[str]:
    if not path.is_file():
        raise HistoricalInputError(f"제외 목록 파일을 찾을 수 없습니다: {path}")
    try:
        with path.open(encoding="utf-8-sig") as file:
            values = json.load(file)
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise HistoricalInputError(f"제외 목록을 읽을 수 없습니다: {path}") from exc
    if not isinstance(values, list):
        raise HistoricalInputError("except_list.json의 최상위 값은 배열이어야 합니다.")
    return {_subcategory_key(value) for value in values if _subcategory_key(value)}


def _find_header(
    rows: Iterable[tuple[object, ...]],
    *,
    sheet_name: str,
) -> tuple[dict[str, int], int, Iterable[tuple[object, ...]]]:
    iterator = iter(rows)
    for row_number, row in enumerate(iterator, start=1):
        names = [_clean_scalar(value) for value in row]
        if ORDER_NUMBER not in names:
            if row_number >= 50:
                break
            continue
        header = {name: index for index, name in enumerate(names) if name}
        missing = sorted(set(TRUTH_COLUMNS) - set(header))
        if missing:
            raise HistoricalInputError(
                f"정답 시트 '{sheet_name}'에 필수 열이 없습니다: {missing}"
            )
        return header, row_number, iterator
    raise HistoricalInputError(
        f"정답 시트 '{sheet_name}'에서 '{ORDER_NUMBER}' 헤더를 찾지 못했습니다."
    )


def _read_truth_workbook(
    truth_path: Path,
    months: tuple[int, ...],
    exclusion_keys: set[str],
    blank_row_limit: int,
) -> tuple[
    dict[str, _TruthRecord],
    dict[int, _TruthSheetSummary],
    int,
    tuple[str, ...],
]:
    if not truth_path.is_file():
        raise HistoricalInputError(f"정답 파일을 찾을 수 없습니다: {truth_path}")

    try:
        workbook = load_workbook(truth_path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl uses several exception types by file state
        raise HistoricalInputError(f"정답 파일을 읽을 수 없습니다: {truth_path}") from exc

    truth_by_order: dict[str, _TruthRecord] = {}
    sheet_summaries: dict[int, _TruthSheetSummary] = {}
    duplicate_orders: list[str] = []
    seen_truth_orders: set[str] = set()
    missing_order_rows = 0
    try:
        sheet_lookup = {_clean_scalar(name): name for name in workbook.sheetnames}
        for month in months:
            requested_name = f"{month}월"
            actual_name = sheet_lookup.get(requested_name)
            if actual_name is None:
                raise HistoricalInputError(
                    f"정답 파일에 필요한 시트가 없습니다: {requested_name}"
                )
            worksheet = workbook[actual_name]
            header, header_row, rows = _find_header(
                worksheet.iter_rows(values_only=True), sheet_name=actual_name
            )

            source_rows = 0
            truth_count = 0
            excluded_department_rows = 0
            excluded_business_rows = 0
            excluded_subcategory_rows = 0
            excluded_cancelled_rows = 0
            sheet_missing_order_rows = 0
            blank_streak = 0

            for row_number, row in enumerate(rows, start=header_row + 1):
                values = {
                    column: row[index] if index < len(row) else None
                    for column, index in header.items()
                    if column in TRUTH_COLUMNS
                }
                order_number = _normalize_order_number(values.get(ORDER_NUMBER))
                if order_number is None:
                    relevant_values = [values.get(column) for column in TRUTH_COLUMNS]
                    if any(_clean_scalar(value) for value in relevant_values):
                        sheet_missing_order_rows += 1
                        missing_order_rows += 1
                        blank_streak = 0
                    else:
                        blank_streak += 1
                        if blank_streak >= blank_row_limit:
                            break
                    continue

                blank_streak = 0
                source_rows += 1
                if order_number in seen_truth_orders:
                    duplicate_orders.append(order_number)
                else:
                    seen_truth_orders.add(order_number)
                department = _clean_scalar(values.get(DEPARTMENT))
                business = _clean_scalar(values.get(BUSINESS))
                subcategory = _clean_scalar(values.get(SUBCATEGORY))
                status = _clean_scalar(values.get(STATUS))
                if department != BUSINESS_DEPARTMENT:
                    excluded_department_rows += 1
                    continue
                if business not in BUSINESS_VALUES:
                    excluded_business_rows += 1
                    continue
                if _subcategory_key(subcategory) in exclusion_keys:
                    excluded_subcategory_rows += 1
                    continue
                if status == CANCELLED_ORDER_STATUS:
                    excluded_cancelled_rows += 1
                    continue

                metric_date = _normalize_date_scalar(values.get(ORDER_DATE))
                person = _clean_scalar(values.get(PERSON))
                missing_dimensions = [
                    name
                    for name, value in (
                        (ORDER_DATE, metric_date),
                        (PERSON, person),
                        (BUSINESS, business),
                        (SUBCATEGORY, subcategory),
                    )
                    if not value
                ]
                if missing_dimensions:
                    raise HistoricalValidationError(
                        f"정답 데이터 결측: {actual_name} 시트 {row_number}행 "
                        f"{order_number}, 열={missing_dimensions}"
                    )
                if int(metric_date[5:7]) != month:
                    raise HistoricalValidationError(
                        f"정답 월 불일치: {actual_name} 시트 {row_number}행 "
                        f"{order_number}, 오더생성일={metric_date}"
                    )

                record = _TruthRecord(
                    order_number=order_number,
                    metric_date=metric_date,
                    person=person,
                    business=business,
                    subcategory=subcategory,
                    month=month,
                    sheet=actual_name,
                    row_number=row_number,
                )
                if order_number not in truth_by_order:
                    truth_by_order[order_number] = record
                truth_count += 1

            sheet_summaries[month] = _TruthSheetSummary(
                source_rows=source_rows,
                truth_count=truth_count,
                excluded_department_rows=excluded_department_rows,
                excluded_business_rows=excluded_business_rows,
                excluded_subcategory_rows=excluded_subcategory_rows,
                excluded_cancelled_rows=excluded_cancelled_rows,
                missing_order_rows=sheet_missing_order_rows,
            )
    finally:
        workbook.close()

    if missing_order_rows:
        raise HistoricalValidationError(
            f"오더번호가 비어 있는 정답 데이터 행이 {missing_order_rows:,}건 있습니다."
        )
    if duplicate_orders:
        sample = ", ".join(sorted(set(duplicate_orders))[:10])
        raise HistoricalValidationError(
            f"중복된 정답 오더번호가 {len(duplicate_orders):,}건 있습니다: {sample}"
        )
    return truth_by_order, sheet_summaries, missing_order_rows, tuple(duplicate_orders)


def _clean_dimension_series(series: pd.Series) -> pd.Series:
    cleaned = series.astype("string").str.strip()
    return cleaned.mask(cleaned.isna() | cleaned.eq(""), UNKNOWN_VALUE)


def _detect_total_header(path: Path) -> int:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise HistoricalInputError(
            f"월별 원본 Excel을 읽을 수 없습니다: {path}"
        ) from exc
    try:
        if not workbook.worksheets:
            raise HistoricalInputError(
                f"월별 원본 Excel에 시트가 없습니다: {path.name}"
            )
        worksheet = workbook.worksheets[0]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=MAX_TOTAL_HEADER_ROWS),
            start=1,
        ):
            values = {
                _clean_scalar(cell.value)
                for cell in row
                if _clean_scalar(cell.value)
            }
            if TOTAL_HEADER_SIGNATURE.issubset(values):
                return row_number - 1
    finally:
        workbook.close()
    raise HistoricalInputError(
        f"월별 원본 Excel의 열 이름 행을 찾지 못했습니다: {path.name}"
    )


def _read_total_source(
    source_path: Path,
    sap_id_map: Mapping[str, str] | None,
) -> pd.DataFrame:
    if source_path.suffix.lower() == ".pkl":
        try:
            raw = pd.read_pickle(source_path)
        except Exception as exc:
            raise HistoricalInputError(
                f"월별 PKL을 읽을 수 없습니다: {source_path}"
            ) from exc
        if not isinstance(raw, pd.DataFrame):
            raise HistoricalInputError(
                f"월별 PKL이 DataFrame이 아닙니다: {source_path}"
            )
        return raw

    if source_path.suffix.lower() != ".xlsx":
        raise HistoricalInputError(
            f"지원하지 않는 월별 원본 형식입니다: {source_path.name}"
        )

    header_row = _detect_total_header(source_path)
    try:
        raw = pd.read_excel(
            source_path,
            sheet_name=0,
            header=header_row,
            usecols=lambda name: str(name).strip() in TOTAL_SOURCE_COLUMNS,
        )
    except Exception as exc:
        raise HistoricalInputError(
            f"월별 원본 Excel을 읽을 수 없습니다: {source_path}"
        ) from exc
    raw.columns = raw.columns.astype(str).str.strip()
    for column in raw.columns:
        if raw[column].dtype == object or isinstance(
            raw[column].dtype, pd.StringDtype
        ):
            raw[column] = raw[column].map(
                lambda value: value.strip() if isinstance(value, str) else value
            )
            raw[column] = raw[column].replace("", pd.NA)
    raw = raw.dropna(axis=0, how="all")
    if ORDER_NUMBER in raw.columns:
        raw = raw.loc[raw[ORDER_NUMBER].notna()].copy()

    required = {
        ORDER_NUMBER,
        ORDER_DATE,
        STATUS,
        SUBCATEGORY,
        ORDER_CREATOR,
        DEPARTMENT,
    }
    missing = sorted(required - set(raw.columns))
    if missing:
        raise HistoricalInputError(
            f"월별 원본 Excel '{source_path.name}'에 필수 열이 없습니다: {missing}"
        )

    creator_id = raw[ORDER_CREATOR].astype("string").str.strip().str.upper()
    normalized_sap = {
        str(key).strip().upper(): str(value).strip()
        for key, value in (sap_id_map or {}).items()
    }
    mapped_person = creator_id.map(normalized_sap).astype("string")
    if PERSON in raw.columns:
        existing_person = raw[PERSON].astype("string").str.strip()
        existing_person = existing_person.mask(existing_person.eq(""))
        # Historical workbooks retain the name at the time of order creation.
        # Current SAP mappings can change later, so they fill blanks only.
        raw[PERSON] = existing_person.fillna(mapped_person)
    else:
        raw[PERSON] = mapped_person

    if CUSTOMER_SERVICE_CENTER in raw.columns:
        center = raw[CUSTOMER_SERVICE_CENTER].astype("string").str.strip()
    elif SERVICE_CENTER in raw.columns:
        center = raw[SERVICE_CENTER].astype("string").str.strip()
    else:
        center = pd.Series(pd.NA, index=raw.index, dtype="string")
    center = center.str.upper().replace({"H051": "H073"})

    creator_business = creator_id.str[:5].map(BUSINESS_BY_CREATOR_PREFIX)
    for creator, business_value in CREATOR_BUSINESS_EXCEPTIONS.items():
        creator_business = creator_business.mask(
            creator_id.eq(creator), business_value
        )
    if BUSINESS in raw.columns:
        existing_business = raw[BUSINESS].astype("string").str.strip()
        existing_business = existing_business.where(
            existing_business.isin(BUSINESS_VALUES)
        )
        creator_business = creator_business.fillna(existing_business)
    raw[BUSINESS] = creator_business.fillna(center.map(BUSINESS_BY_CENTER))

    return raw.reset_index(drop=True)


def _reduce_month(
    source_path: Path,
    month: int,
    exclusion_keys: set[str],
    truth_by_order: Mapping[str, _TruthRecord],
    truth_match_counts: dict[str, int],
    sap_id_map: Mapping[str, str] | None = None,
) -> tuple[pd.DataFrame, dict[str, int], list[str]]:
    raw = _read_total_source(source_path, sap_id_map)

    raw.columns = raw.columns.astype(str).str.strip()
    missing = sorted(set(TOTAL_COLUMNS) - set(raw.columns))
    if missing:
        raise HistoricalInputError(
            f"월별 원본 '{source_path.name}'에 필수 열이 없습니다: {missing}"
        )

    source_rows = len(raw)
    frame = raw.loc[:, list(TOTAL_COLUMNS)].copy()
    del raw

    department = frame[DEPARTMENT].astype("string").str.strip()
    business = frame[BUSINESS].astype("string").str.strip()
    subcategory = frame[SUBCATEGORY].astype("string").str.strip()
    status = frame[STATUS].astype("string").str.strip()
    department_mask = department.eq(BUSINESS_DEPARTMENT).fillna(False)
    business_mask = business.isin(BUSINESS_VALUES).fillna(False)
    subcategory_mask = ~subcategory.map(_subcategory_key).isin(exclusion_keys)
    cancelled_mask = status.eq(CANCELLED_ORDER_STATUS).fillna(False)
    # Dashboard denominators intentionally include except_list subcategories.
    # The exclusion list still limits confirmed-error/candidate classification,
    # while displayed totals use every business-department, non-cancelled row.
    keep_mask = department_mask & business_mask & ~cancelled_mask

    excluded_department_rows = int((~department_mask).sum())
    excluded_business_rows = int(
        (department_mask & ~business_mask).sum()
    )
    excluded_subcategory_rows = int(
        (department_mask & business_mask & ~subcategory_mask).sum()
    )
    excluded_cancelled_rows = int(
        (department_mask & business_mask & cancelled_mask).sum()
    )
    frame = frame.loc[keep_mask].copy()
    if frame.empty:
        raise HistoricalValidationError(
            f"현재 집계 기준 적용 후 데이터가 없습니다: {source_path.name}"
        )

    frame["_order_number"] = _normalize_order_series(frame[ORDER_NUMBER])
    missing_total_orders = int(frame["_order_number"].isna().sum())
    if missing_total_orders:
        raise HistoricalValidationError(
            f"'{source_path.name}'에 오더번호 결측이 {missing_total_orders:,}건 있습니다."
        )

    converted_dates = pd.to_datetime(frame[ORDER_DATE], errors="coerce")
    invalid_dates = int(converted_dates.isna().sum())
    if invalid_dates:
        raise HistoricalValidationError(
            f"'{source_path.name}'에 변환할 수 없는 오더생성일이 {invalid_dates:,}건 있습니다."
        )
    frame["metric_date"] = converted_dates.dt.strftime("%Y-%m-%d")
    wrong_month = ~converted_dates.dt.month.eq(month)
    if wrong_month.any():
        sample = frame.loc[wrong_month, [ORDER_NUMBER, "metric_date"]].head(5)
        raise HistoricalValidationError(
            f"'{source_path.name}'에 다른 월 데이터가 {int(wrong_month.sum()):,}건 있습니다: "
            f"{sample.to_dict('records')}"
        )

    frame["person"] = _clean_dimension_series(frame[PERSON])
    frame["business"] = _clean_dimension_series(frame[BUSINESS])
    frame["subcategory"] = _clean_dimension_series(frame[SUBCATEGORY])

    truth_keys = set(truth_by_order)
    frame["_is_error"] = frame["_order_number"].isin(truth_keys)
    matched = frame.loc[
        frame["_is_error"],
        ["_order_number", "metric_date", "person", "business", "subcategory"],
    ]
    mismatches: list[str] = []
    for row in matched.itertuples(index=False, name=None):
        order_number, metric_date, person, business, subcategory_value = row
        order_number = str(order_number)
        truth_match_counts[order_number] += 1
        truth = truth_by_order[order_number]
        actual = (
            str(metric_date),
            str(person),
            str(business),
            str(subcategory_value),
        )
        expected = (
            truth.metric_date,
            truth.person,
            truth.business,
            truth.subcategory,
        )
        if actual != expected:
            labels = (ORDER_DATE, PERSON, BUSINESS, SUBCATEGORY)
            differences = ", ".join(
                f"{label}: 정답={want!r}/전체={got!r}"
                for label, want, got in zip(labels, expected, actual)
                if want != got
            )
            mismatches.append(f"{order_number} ({differences})")

    aggregate = (
        frame.groupby(
            ["metric_date", "person", "business", "subcategory"],
            dropna=False,
            observed=True,
        )
        .agg(total_count=("_order_number", "size"), error_count=("_is_error", "sum"))
        .reset_index()
    )
    aggregate["total_count"] = aggregate["total_count"].astype("int64")
    aggregate["error_count"] = aggregate["error_count"].astype("int64")

    counts = {
        "source_rows": source_rows,
        "total_count": len(frame),
        "excluded_department_rows": excluded_department_rows,
        "excluded_business_rows": excluded_business_rows,
        "excluded_subcategory_rows": excluded_subcategory_rows,
        "excluded_cancelled_rows": excluded_cancelled_rows,
        "matched_truth_count": int(frame["_is_error"].sum()),
        "aggregate_rows": len(aggregate),
    }
    return aggregate.loc[:, list(AGGREGATE_COLUMNS)], counts, mismatches


def load_historical_aggregates(
    source_dir: str | Path,
    truth_path: str | Path,
    except_list_path: str | Path,
    *,
    months: Iterable[int] = range(1, 7),
    blank_row_limit: int = 200,
    sap_id_map: Mapping[str, str] | None = None,
) -> HistoricalLoadResult:
    """Validate and reduce historical totals and confirmed errors.

    Parameters
    ----------
    source_dir:
        Directory containing the original monthly ``26년 N월 전체 서비스오더
        리스트.xlsx`` files. Legacy ``N월.pkl`` files are accepted only when
        the corresponding original workbook is absent.
    truth_path:
        Workbook whose ``1월`` through ``6월`` sheets contain confirmed errors.
    except_list_path:
        JSON list used to limit confirmed-error scope. These subcategories
        remain in the displayed dashboard denominator.
    months:
        Month numbers to load.  Defaults to January through June.
    blank_row_limit:
        Stop streaming a truth sheet after this many consecutive empty rows.

    Returns
    -------
    HistoricalLoadResult
        ``aggregates`` contains only ``metric_date``, ``person``, ``business``,
        ``subcategory``, ``total_count``, and ``error_count``.  No order-level
        truth or total rows are retained in the result or written to disk.

    Raises
    ------
    HistoricalInputError
        If required files, sheets, or columns are absent.
    HistoricalValidationError
        If truth IDs are missing/duplicated/unmatched, or if truth dimensions
        differ from the corresponding total-data rows.
    """

    month_values = tuple(int(month) for month in months)
    if not month_values or len(set(month_values)) != len(month_values):
        raise ValueError("months에는 중복 없는 월 번호가 하나 이상 필요합니다.")
    if any(month < 1 or month > 12 for month in month_values):
        raise ValueError("월 번호는 1부터 12 사이여야 합니다.")
    if blank_row_limit < 1:
        raise ValueError("blank_row_limit은 1 이상이어야 합니다.")

    source_directory = Path(source_dir)
    truth_file = Path(truth_path)
    if not source_directory.is_dir():
        raise HistoricalInputError(
            f"월별 원본 폴더를 찾을 수 없습니다: {source_directory}"
        )

    exclusion_keys = _load_exclusion_keys(Path(except_list_path))
    truth_by_order, truth_sheets, missing_truth_orders, duplicate_truth_orders = (
        _read_truth_workbook(
            truth_file,
            month_values,
            exclusion_keys,
            blank_row_limit,
        )
    )
    truth_match_counts = {order_number: 0 for order_number in truth_by_order}

    monthly_aggregates: list[pd.DataFrame] = []
    monthly_summaries: list[MonthLoadSummary] = []
    all_mismatches: list[str] = []

    for month in month_values:
        original_path = (
            source_directory / f"26년 {month}월 전체 서비스오더 리스트.xlsx"
        )
        legacy_path = source_directory / f"{month}월.pkl"
        source_path = original_path if original_path.is_file() else legacy_path
        if not source_path.is_file():
            raise HistoricalInputError(
                "월별 원본을 찾을 수 없습니다: "
                f"{original_path.name} 또는 {legacy_path.name}"
            )
        aggregate, counts, mismatches = _reduce_month(
            source_path,
            month,
            exclusion_keys,
            truth_by_order,
            truth_match_counts,
            sap_id_map,
        )
        monthly_aggregates.append(aggregate)
        all_mismatches.extend(mismatches)
        truth_sheet = truth_sheets[month]
        monthly_summaries.append(
            MonthLoadSummary(
                month=month,
                source_rows=counts["source_rows"],
                total_count=counts["total_count"],
                excluded_department_rows=counts["excluded_department_rows"],
                excluded_business_rows=counts["excluded_business_rows"],
                excluded_subcategory_rows=counts["excluded_subcategory_rows"],
                excluded_cancelled_rows=counts["excluded_cancelled_rows"],
                truth_source_rows=truth_sheet.source_rows,
                truth_count=truth_sheet.truth_count,
                truth_excluded_department_rows=truth_sheet.excluded_department_rows,
                truth_excluded_business_rows=truth_sheet.excluded_business_rows,
                truth_excluded_subcategory_rows=truth_sheet.excluded_subcategory_rows,
                truth_excluded_cancelled_rows=truth_sheet.excluded_cancelled_rows,
                matched_truth_count=counts["matched_truth_count"],
                aggregate_rows=counts["aggregate_rows"],
            )
        )

    unmatched = sorted(
        order_number
        for order_number, match_count in truth_match_counts.items()
        if match_count == 0
    )
    multiply_matched = sorted(
        order_number
        for order_number, match_count in truth_match_counts.items()
        if match_count > 1
    )
    # Confirmed-error rows that do not exist in the defined total-data scope
    # cannot form a valid numerator/denominator pair. Record them in the load
    # summary, but do not inject synthetic total rows or abort the remaining
    # valid aggregation.
    validation_messages: list[str] = []
    if multiply_matched:
        validation_messages.append(
            "전체 데이터에서 두 번 이상 매칭된 정답 "
            f"{len(multiply_matched):,}건: {', '.join(multiply_matched[:10])}"
        )
    if all_mismatches:
        validation_messages.append(
            f"정답 차원 불일치 {len(all_mismatches):,}건: "
            + "; ".join(all_mismatches[:10])
        )
    if validation_messages:
        raise HistoricalValidationError(" | ".join(validation_messages))

    combined = pd.concat(monthly_aggregates, ignore_index=True)
    aggregates = (
        combined.groupby(
            ["metric_date", "person", "business", "subcategory"],
            dropna=False,
            observed=True,
            as_index=False,
        )[["total_count", "error_count"]]
        .sum()
        .sort_values(
            ["metric_date", "business", "person", "subcategory"],
            kind="stable",
        )
        .reset_index(drop=True)
    )
    aggregates["total_count"] = aggregates["total_count"].astype("int64")
    aggregates["error_count"] = aggregates["error_count"].astype("int64")
    aggregates = aggregates.loc[:, list(AGGREGATE_COLUMNS)]

    summary = HistoricalLoadSummary(
        months=month_values,
        source_total_rows=sum(item.source_rows for item in monthly_summaries),
        total_count=int(aggregates["total_count"].sum()),
        excluded_department_rows=sum(
            item.excluded_department_rows for item in monthly_summaries
        ),
        excluded_business_rows=sum(
            item.excluded_business_rows for item in monthly_summaries
        ),
        excluded_subcategory_rows=sum(
            item.excluded_subcategory_rows for item in monthly_summaries
        ),
        excluded_cancelled_rows=sum(
            item.excluded_cancelled_rows for item in monthly_summaries
        ),
        truth_source_rows=sum(item.truth_source_rows for item in monthly_summaries),
        truth_count=int(aggregates["error_count"].sum()),
        truth_excluded_department_rows=sum(
            item.truth_excluded_department_rows for item in monthly_summaries
        ),
        truth_excluded_business_rows=sum(
            item.truth_excluded_business_rows for item in monthly_summaries
        ),
        truth_excluded_subcategory_rows=sum(
            item.truth_excluded_subcategory_rows for item in monthly_summaries
        ),
        truth_excluded_cancelled_rows=sum(
            item.truth_excluded_cancelled_rows for item in monthly_summaries
        ),
        truth_missing_order_rows=missing_truth_orders,
        truth_duplicate_order_count=len(duplicate_truth_orders),
        truth_unmatched_count=len(unmatched),
        truth_dimension_mismatch_count=len(all_mismatches),
        aggregate_rows=len(aggregates),
        month_summaries=tuple(monthly_summaries),
    )
    return HistoricalLoadResult(aggregates=aggregates, summary=summary)


__all__ = [
    "AGGREGATE_COLUMNS",
    "HistoricalDataError",
    "HistoricalInputError",
    "HistoricalLoadResult",
    "HistoricalLoadSummary",
    "HistoricalValidationError",
    "MonthLoadSummary",
    "load_historical_aggregates",
]
