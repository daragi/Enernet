from __future__ import annotations

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase, main
from unittest.mock import patch
import json
import sqlite3
import sys

import pandas as pd


CENTER_DIR = Path(__file__).resolve().parents[1]
if str(CENTER_DIR) not in sys.path:
    sys.path.insert(0, str(CENTER_DIR))

from service_order.error import analysis_pipeline  # noqa: E402
from service_order.service_order_store import DashboardStore  # noqa: E402


def order_frame(order_numbers: list[str]) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "오더번호": order_numbers,
            "오더생성일": [pd.Timestamp("2026-07-08")] * len(order_numbers),
            "생성인": ["테스트"] * len(order_numbers),
            "사업부": ["중부"] * len(order_numbers),
            "소분류": ["공급중지"] * len(order_numbers),
            "내역": [f"상세 {value}" for value in order_numbers],
        }
    )


class ExactClassificationTest(TestCase):
    def test_dominant_other_subcategory_phrase_keeps_anomaly_for_review(self) -> None:
        source = order_frame(["anomaly", "owner-normal"])
        source["소분류"] = ["공급중지", "체납"]
        source["내역"] = [
            "공급중지 예고장 송달 까지",
            "공급중지 예고장 송달 까지",
        ]
        with TemporaryDirectory() as directory:
            root = Path(directory)
            keyword = root / "keyword.json"
            auto_normal = root / "auto_normal_pattern.json"
            empty = root / "empty.json"
            keyword.write_text(
                json.dumps(
                    {
                        "공급중지": {"단일키워드": ["공급중지"]},
                        "체납": {"단일키워드": ["예고장"]},
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            auto_normal.write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "source_subcategory": "체납",
                                "pattern_type": "4어절문구",
                                "pattern": "공급중지 예고장 송달 까지",
                                "status": "proposed",
                                "normal_support": 20,
                                "other_subcategory_frequency": 0,
                                "confirmed_error_pattern_hits": 0,
                            },
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            empty.write_text("{}", encoding="utf-8")
            with (
                patch.object(analysis_pipeline, "KEYWORD_PATH", keyword),
                patch.object(analysis_pipeline, "CONFLICT_KEYWORD_PATH", empty),
                patch.object(analysis_pipeline, "PRIORITY_KEYWORD_PATH", keyword),
                patch.object(
                    analysis_pipeline,
                    "ERROR_PATTERN_PATH",
                    root / "missing_error.json",
                ),
                patch.object(
                    analysis_pipeline,
                    "AUTO_NORMAL_PATTERN_PATH",
                    auto_normal,
                ),
                patch.object(
                    analysis_pipeline,
                    "SCOPED_NORMAL_PATTERN_PATH",
                    root / "missing_scoped.json",
                ),
                patch.object(
                    analysis_pipeline,
                    "ERROR_LEARNING_BASELINE_PATH",
                    root / "missing_baseline.sqlite3",
                ),
            ):
                candidates, auto_errors, summary = (
                    analysis_pipeline.select_candidate_orders(source, {})
                )

        self.assertTrue(auto_errors.empty)
        self.assertEqual(candidates["오더번호"].tolist(), ["anomaly"])
        self.assertEqual(summary["소분류지배문구이상행수"], 1)

    def test_similarity_promotes_only_when_historical_normal_is_far(self) -> None:
        source = order_frame(["similar-error", "normal-near", "other-owner"])
        source["소분류"] = ["계량기일반", "계량기일반", "검침"]
        source["내역"] = [
            "혹서기 격월 검침 방문",
            "계량기사진 부탁드립니다",
            "혹서기 격월 검침 방문",
        ]
        with TemporaryDirectory() as directory:
            root = Path(directory)
            registry = root / "error_pattern.json"
            empty = root / "empty.json"
            baseline = root / "baseline.sqlite3"
            registry.write_text(
                json.dumps(
                    {
                        "version": 2,
                        "records": [
                            {
                                "source_subcategory": "계량기일반",
                                "signature": "혹서기 격월검침 문의",
                                "status": "active",
                            },
                            {
                                "source_subcategory": "계량기일반",
                                "signature": "해당건물 층 계량기사진 부탁드립니다",
                                "status": "active",
                            },
                        ],
                        "phrase_records": [],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            empty.write_text("{}", encoding="utf-8")
            connection = sqlite3.connect(baseline)
            try:
                connection.executescript(
                    """
                    CREATE TABLE signature_counts (
                        source_subcategory TEXT NOT NULL,
                        signature TEXT NOT NULL,
                        row_count INTEGER NOT NULL
                    );
                    CREATE TABLE truth_sources (
                        order_number TEXT PRIMARY KEY,
                        source_subcategory TEXT NOT NULL,
                        signature TEXT NOT NULL
                    );
                    """
                )
                connection.executemany(
                    """
                    INSERT INTO signature_counts(
                        source_subcategory, signature, row_count
                    ) VALUES (?, ?, ?)
                    """,
                    [
                        ("계량기일반", "보정기 고장 검침방", 1),
                        ("계량기일반", "계량기사진 부탁드립니다", 4),
                    ],
                )
                connection.commit()
            finally:
                connection.close()
            with (
                patch.object(analysis_pipeline, "ERROR_PATTERN_PATH", registry),
                patch.object(analysis_pipeline, "KEYWORD_PATH", empty),
                patch.object(analysis_pipeline, "CONFLICT_KEYWORD_PATH", empty),
                patch.object(analysis_pipeline, "PRIORITY_KEYWORD_PATH", empty),
                patch.object(
                    analysis_pipeline,
                    "AUTO_NORMAL_PATTERN_PATH",
                    root / "missing_auto.json",
                ),
                patch.object(
                    analysis_pipeline,
                    "SCOPED_NORMAL_PATTERN_PATH",
                    root / "missing_scoped.json",
                ),
                patch.object(
                    analysis_pipeline,
                    "ERROR_LEARNING_BASELINE_PATH",
                    baseline,
                ),
            ):
                candidates, auto_errors, summary = (
                    analysis_pipeline.select_candidate_orders(source, {})
                )

        self.assertEqual(auto_errors["오더번호"].tolist(), ["similar-error"])
        self.assertEqual(
            set(candidates["오더번호"].tolist()),
            {"normal-near", "other-owner"},
        )
        self.assertEqual(summary["유사문장자동오생성행수"], 1)

    def test_similarity_template_removes_phone_and_following_name(self) -> None:
        first = analysis_pipeline.normalize_error_similarity_template(
            "계량기지침확인필요 010-1234-5678 김봉기"
        )
        second = analysis_pipeline.normalize_error_similarity_template(
            "계량기지침확인필요 010-9999-0000 이동명"
        )

        self.assertEqual(first, "계량기지침확인필요")
        self.assertEqual(first, second)

    def test_similarity_template_generalizes_amount_date_and_period(self) -> None:
        first = analysis_pipeline.normalize_error_similarity_template(
            "58,404원 24일까지 2개월 체납"
        )
        second = analysis_pipeline.normalize_error_similarity_template(
            "12,000원 30일까지 3개월 체납"
        )

        self.assertEqual(first, "금액값 날짜값 기간값 체납")
        self.assertEqual(first, second)

    def test_auto_normal_evidence_replaces_overlapping_upload(self) -> None:
        source = order_frame([f"N{index}" for index in range(12)])
        source["오더생성일"] = [
            pd.Timestamp(f"2026-07-{(index % 4) + 1:02d}")
            for index in range(12)
        ]
        source["생성인"] = [f"담당{index % 4}" for index in range(12)]
        source["내역"] = ["누적 정상 안전문구 예시"] * 12
        candidates = source.iloc[[0]].copy()
        auto_errors = source.iloc[0:0].copy()
        with TemporaryDirectory() as directory:
            root = Path(directory)
            registry = root / "auto_normal_pattern.json"
            with (
                patch.object(
                    analysis_pipeline, "AUTO_NORMAL_PATTERN_PATH", registry
                ),
                patch.object(
                    analysis_pipeline,
                    "ERROR_PATTERN_PATH",
                    root / "missing_error_pattern.json",
                ),
            ):
                first = analysis_pipeline.update_auto_normal_pattern_registry(
                    source,
                    candidates,
                    auto_errors,
                    {},
                )
                first_document = json.loads(registry.read_text(encoding="utf-8"))
                second = analysis_pipeline.update_auto_normal_pattern_registry(
                    source,
                    candidates,
                    auto_errors,
                    {},
                )
                second_document = json.loads(registry.read_text(encoding="utf-8"))

        self.assertEqual(first["evidence_snapshot_total"], 1)
        self.assertEqual(second["evidence_snapshot_total"], 1)
        self.assertEqual(len(second_document["evidence_snapshots"]), 1)
        first_support = max(
            record["normal_support"] for record in first_document["records"]
        )
        second_support = max(
            record["normal_support"] for record in second_document["records"]
        )
        self.assertEqual(first_support, second_support)

    def test_two_token_normal_phrase_is_promoted_with_diverse_evidence(self) -> None:
        source = order_frame([str(1000 + index) for index in range(11)])
        source.iloc[:, 1] = [
            pd.Timestamp(f"2026-07-{(index % 3) + 1:02d}")
            for index in range(11)
        ]
        source.iloc[:, 2] = [f"creator-{index % 2}" for index in range(11)]
        source.iloc[:, 5] = ["stable phrase"] * 11
        candidates = source.iloc[[0]].copy()
        auto_errors = source.iloc[0:0].copy()
        with TemporaryDirectory() as directory:
            registry = Path(directory) / "auto_normal_pattern.json"
            with (
                patch.object(
                    analysis_pipeline, "AUTO_NORMAL_PATTERN_PATH", registry
                ),
                patch.object(
                    analysis_pipeline,
                    "ERROR_PATTERN_PATH",
                    Path(directory) / "missing_error_pattern.json",
                ),
            ):
                summary = analysis_pipeline.update_auto_normal_pattern_registry(
                    source,
                    candidates,
                    auto_errors,
                    {},
                )
                document = json.loads(registry.read_text(encoding="utf-8"))

        record = next(
            item for item in document["records"] if item["pattern"] == "stable phrase"
        )
        self.assertEqual(record["status"], "active")
        self.assertEqual(summary["new_active"], 1)
        self.assertEqual(summary["current_candidate_impact"], 1)

    def test_repeated_owner_only_candidate_phrase_is_auto_promoted(self) -> None:
        source = order_frame([f"R{index}" for index in range(30)])
        source["오더생성일"] = [
            pd.Timestamp(f"2026-07-{(index % 3) + 1:02d}")
            for index in range(30)
        ]
        source["생성인"] = [f"담당{index % 2}" for index in range(30)]
        source["내역"] = ["반복 정상 업무 문구"] * 30
        candidates = source.iloc[:9].copy()
        auto_errors = source.iloc[0:0].copy()
        with TemporaryDirectory() as directory:
            registry = Path(directory) / "auto_normal_pattern.json"
            with (
                patch.object(
                    analysis_pipeline, "AUTO_NORMAL_PATTERN_PATH", registry
                ),
                patch.object(
                    analysis_pipeline,
                    "ERROR_PATTERN_PATH",
                    Path(directory) / "missing_error_pattern.json",
                ),
            ):
                summary = analysis_pipeline.update_auto_normal_pattern_registry(
                    source,
                    candidates,
                    auto_errors,
                    {},
                )
                document = json.loads(registry.read_text(encoding="utf-8"))

        record = next(
            item
            for item in document["records"]
            if item["pattern"] == "반복 정상 업무 문구"
        )
        self.assertEqual(record["status"], "active")
        self.assertEqual(summary["repeated_candidate_promotions"], 1)
        self.assertEqual(summary["current_candidate_impact"], 9)

    def test_active_is_auto_and_ambiguous_stays_review(self) -> None:
        source = order_frame(["A", "B"])
        source["내역"] = ["같은 확정 문장", "검토 유지 문장"]
        with TemporaryDirectory() as directory:
            registry = Path(directory) / "error_pattern.json"
            registry.write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "source_subcategory": "공급중지",
                                "signature": "같은 확정 문장",
                                "status": "active",
                            },
                            {
                                "source_subcategory": "공급중지",
                                "signature": "검토 유지 문장",
                                "status": "ambiguous",
                            },
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            with patch.object(analysis_pipeline, "ERROR_PATTERN_PATH", registry):
                with patch.object(
                    analysis_pipeline,
                    "AUTO_NORMAL_PATTERN_PATH",
                    root := Path(directory) / "missing_auto_normal.json",
                ):
                    with patch.object(
                        analysis_pipeline,
                        "SCOPED_NORMAL_PATTERN_PATH",
                        root / "missing_scoped_normal.json",
                    ):
                        review, auto, summary = (
                            analysis_pipeline.select_candidate_orders(source, {})
                        )

        self.assertEqual(auto["오더번호"].tolist(), ["A"])
        self.assertEqual(review["오더번호"].tolist(), ["B"])
        self.assertEqual(summary["자동오생성행수"], 1)
        self.assertEqual(summary["검토후보행수"], 1)

    def test_soft_single_priority_and_ambiguous_review_order(self) -> None:
        source = order_frame(["soft", "hard", "owner", "ambiguous"])
        source["소분류"] = ["검침", "검침", "서비스기타", "검침"]
        source["내역"] = [
            "검침 혹서기",
            "검침 혹서기 문의",
            "서비스 검침",
            "검침",
        ]
        with TemporaryDirectory() as directory:
            root = Path(directory)
            keyword = root / "keyword.json"
            conflict = root / "conflict_keyword.json"
            priority = root / "priority_keyword.json"
            registry = root / "error_pattern.json"
            keyword.write_text(
                json.dumps(
                    {
                        "검침": {"단일키워드": ["검침"]},
                        "서비스기타": {"단일키워드": ["서비스"]},
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            conflict.write_text(
                json.dumps(
                    {
                        "서비스기타": {
                            "단일키워드": ["혹서기"],
                            "2어절문구": ["혹서기 문의"],
                        }
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            priority.write_text(
                json.dumps(
                    {"검침": {"단일키워드": ["검침"]}},
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            registry.write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "source_subcategory": "검침",
                                "signature": "검침",
                                "status": "ambiguous",
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            with (
                patch.object(analysis_pipeline, "KEYWORD_PATH", keyword),
                patch.object(analysis_pipeline, "CONFLICT_KEYWORD_PATH", conflict),
                patch.object(analysis_pipeline, "PRIORITY_KEYWORD_PATH", priority),
                patch.object(analysis_pipeline, "ERROR_PATTERN_PATH", registry),
                patch.object(
                    analysis_pipeline,
                    "AUTO_NORMAL_PATTERN_PATH",
                    root / "missing_auto_normal.json",
                ),
                patch.object(
                    analysis_pipeline,
                    "SCOPED_NORMAL_PATTERN_PATH",
                    root / "missing_scoped_normal.json",
                ),
            ):
                review, auto, summary = analysis_pipeline.select_candidate_orders(
                    source, {}
                )

        self.assertTrue(auto.empty)
        self.assertEqual(
            set(review["오더번호"].tolist()), {"hard", "ambiguous"}
        )
        self.assertEqual(summary["일반단일충돌약화행수"], 2)
        self.assertEqual(summary["확정검토보호행수"], 1)

    def test_scoped_normal_overrides_collision_but_not_error_or_review(self) -> None:
        source = order_frame(["normal", "error", "review"])
        source["소분류"] = ["검침", "검침", "검침"]
        source["내역"] = [
            "정상전용 위험문구",
            "확정 정상전용 위험문구",
            "검토 정상전용 위험문구",
        ]
        with TemporaryDirectory() as directory:
            root = Path(directory)
            keyword = root / "keyword.json"
            conflict = root / "conflict_keyword.json"
            priority = root / "priority_keyword.json"
            registry = root / "error_pattern.json"
            scoped = root / "scoped_normal_pattern.json"
            keyword.write_text(
                json.dumps({"검침": {"단일키워드": ["정상전용"]}}, ensure_ascii=False),
                encoding="utf-8",
            )
            conflict.write_text(
                json.dumps(
                    {"서비스기타": {"2어절문구": ["정상전용 위험문구"]}},
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            priority.write_text("{}", encoding="utf-8")
            registry.write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "source_subcategory": "검침",
                                "signature": "확정 정상전용 위험문구",
                                "status": "active",
                            },
                            {
                                "source_subcategory": "검침",
                                "signature": "검토 정상전용 위험문구",
                                "status": "ambiguous",
                            },
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            scoped.write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "source_subcategory": "검침",
                                "pattern_type": "2어절문구",
                                "pattern": "정상전용 위험문구",
                                "status": "active",
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            with (
                patch.object(analysis_pipeline, "KEYWORD_PATH", keyword),
                patch.object(analysis_pipeline, "CONFLICT_KEYWORD_PATH", conflict),
                patch.object(analysis_pipeline, "PRIORITY_KEYWORD_PATH", priority),
                patch.object(analysis_pipeline, "ERROR_PATTERN_PATH", registry),
                patch.object(
                    analysis_pipeline,
                    "AUTO_NORMAL_PATTERN_PATH",
                    root / "missing_auto_normal.json",
                ),
                patch.object(analysis_pipeline, "SCOPED_NORMAL_PATTERN_PATH", scoped),
            ):
                review, auto, summary = analysis_pipeline.select_candidate_orders(
                    source, {}
                )

        self.assertEqual(review["오더번호"].tolist(), ["review"])
        self.assertEqual(auto["오더번호"].tolist(), ["error"])
        self.assertEqual(summary["소분류확정정상충돌우선행수"], 1)
        self.assertEqual(summary["확정검토보호행수"], 1)

    def test_classification_workbook_contains_review_and_auto_sheets(self) -> None:
        review = order_frame(["R"])
        auto = order_frame(["A"])
        with TemporaryDirectory() as directory:
            path = Path(directory) / "combined.xlsx"
            analysis_pipeline.save_classification_workbook(review, auto, path)
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=False, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["후보군", "자동오생성"])
                self.assertEqual(workbook["후보군"].max_row - 1, 1)
                self.assertEqual(workbook["자동오생성"].max_row - 1, 1)
                self.assertTrue(workbook["후보군"].sheet_view.showGridLines)
                self.assertTrue(workbook["자동오생성"].sheet_view.showGridLines)
            finally:
                workbook.close()


class AutomaticPersistenceTest(TestCase):
    def test_confirmed_error_exclusion_survives_reupload_and_restores(self) -> None:
        with TemporaryDirectory() as directory:
            store = DashboardStore(Path(directory) / "metrics.sqlite3")
            totals = order_frame(["A", "B"])
            now = datetime.now().astimezone().isoformat(timespec="microseconds")
            store.replace_totals_and_auto_errors(
                totals,
                totals,
                batch_id="auto-first",
                job_id="job-one",
                source_name="one.xlsx",
                approved_at=now,
            )
            excluded = store.exclude_error_orders(
                ["A"],
                job_id="job-one",
                excluded_at=now,
            )
            self.assertEqual(excluded["excluded_count"], 1)
            self.assertEqual(store.status()["error_count"], 1)
            self.assertEqual(store.excluded_error_order_numbers(), {"A"})

            store.replace_totals_and_auto_errors(
                totals,
                totals,
                batch_id="auto-second",
                job_id="job-two",
                source_name="two.xlsx",
                approved_at=now,
            )
            self.assertEqual(store.status()["error_count"], 1)

            restored = store.restore_error_orders(
                ["A"],
                job_id="job-two",
                restored_at=now,
            )
            self.assertEqual(restored["restored_count"], 1)
            self.assertEqual(store.status()["error_count"], 2)
            self.assertEqual(store.excluded_error_order_numbers(), set())

    def test_current_job_refresh_persists_reclassified_workbook(self) -> None:
        import dashboard_server as server

        with TemporaryDirectory() as directory:
            root = Path(directory)
            results = root / "results"
            job_id = "abcdef123456"
            job_dir = results / job_id
            job_dir.mkdir(parents=True)
            candidate_file = "7_1-8_candidate.xlsx"
            candidate_path = job_dir / candidate_file
            source = order_frame(["A", "B"])
            analysis_pipeline.save_classification_workbook(
                source,
                source.iloc[0:0],
                candidate_path,
            )
            store = DashboardStore(root / "metrics.sqlite3")
            store.replace_totals(source)
            manager = server.CurrentJobManager()
            with (
                patch.object(server, "DASHBOARD_STORE", store),
                patch.object(server, "WEB_OUTPUT_ROOT", results),
                patch.object(
                    server,
                    "CURRENT_JOB_MANIFEST_PATH",
                    results / "current_job.json",
                ),
            ):
                manager.replace(
                    job_id,
                    "source.xlsx",
                    source,
                    source,
                    period={"start": "2026-07-08", "end": "2026-07-08"},
                    candidate_summary={"자동정상누적기간수": 1},
                    preprocessed_file="source_total_data.xlsx",
                    candidate_file=candidate_file,
                )
                with (
                    patch.object(server, "load_configuration", return_value=({}, set())),
                    patch.object(
                        server,
                        "select_candidate_orders",
                        return_value=(
                            source.iloc[[1]].reset_index(drop=True),
                            source.iloc[[0]].reset_index(drop=True),
                            {
                                "후보행수": 1,
                                "자동오생성행수": 1,
                                "정상제외행수": 0,
                            },
                        ),
                    ),
                ):
                    refreshed = manager.refresh_classification(job_id)

            from openpyxl import load_workbook

            workbook = load_workbook(candidate_path, read_only=True)
            try:
                self.assertEqual(workbook["후보군"].max_row - 1, 1)
                self.assertEqual(workbook["자동오생성"].max_row - 1, 1)
            finally:
                workbook.close()
            self.assertEqual(refreshed["candidate_count"], 1)
            self.assertEqual(refreshed["auto_error_count"], 1)
            self.assertEqual(
                refreshed["candidate"]["자동정상누적기간수"], 1
            )

    def test_manual_approval_moves_from_candidate_to_confirmed_grid(self) -> None:
        import dashboard_server as server

        with TemporaryDirectory() as directory:
            store = DashboardStore(Path(directory) / "metrics.sqlite3")
            source = order_frame(["A", "B"])
            store.replace_totals(source)
            now = datetime.now().astimezone().isoformat(timespec="microseconds")
            store.approve_error_batch(
                batch_id="manual-grid",
                job_id="job-grid",
                source_name="source.xlsx",
                approved_at=now,
                records=[
                    {
                        "candidate_row_id": 0,
                        "order_number": "A",
                        "order_date": "2026-07-08",
                        "person": "테스트",
                        "business": "중부",
                        "subcategory": "공급중지",
                        "payload": {"오더번호": "A"},
                    }
                ],
            )
            with patch.object(server, "DASHBOARD_STORE", store):
                candidates, confirmed, view, excluded = (
                    server.apply_error_exclusions(
                        source,
                        source,
                        source.iloc[0:0],
                    )
                )

        self.assertEqual(candidates["오더번호"].tolist(), ["B"])
        self.assertEqual(confirmed["오더번호"].tolist(), ["A"])
        self.assertEqual(view["집계상태"].tolist(), ["확정"])
        self.assertEqual(excluded, 0)

    def test_reupload_replaces_auto_preserves_manual_and_rolls_back_manual(self) -> None:
        with TemporaryDirectory() as directory:
            store = DashboardStore(Path(directory) / "metrics.sqlite3")
            totals = order_frame(["A", "B", "C"])
            now = datetime.now().astimezone().isoformat(timespec="microseconds")
            first = store.replace_totals_and_auto_errors(
                totals,
                totals.iloc[[0]],
                batch_id="auto-one",
                job_id="job-one",
                source_name="one.xlsx",
                approved_at=now,
            )
            self.assertEqual(first["auto_error_count"], 1)

            store.approve_error_batch(
                batch_id="manual-one",
                job_id="job-one",
                source_name="one.xlsx",
                approved_at=now,
                records=[
                    {
                        "candidate_row_id": 1,
                        "order_number": "B",
                        "order_date": "2026-07-08",
                        "person": "테스트",
                        "business": "중부",
                        "subcategory": "공급중지",
                        "payload": {"오더번호": "B"},
                    }
                ],
            )

            second = store.replace_totals_and_auto_errors(
                totals,
                totals.iloc[[0, 1]],
                batch_id="auto-two",
                job_id="job-two",
                source_name="two.xlsx",
                approved_at=now,
            )
            self.assertEqual(second["auto_replaced_count"], 1)
            self.assertEqual(second["auto_error_count"], 1)
            self.assertEqual(second["auto_skipped_manual_count"], 1)
            self.assertEqual(second["error_count"], 2)

            third = store.replace_totals_and_auto_errors(
                totals,
                totals.iloc[[2]],
                batch_id="auto-three",
                job_id="job-three",
                source_name="three.xlsx",
                approved_at=now,
            )
            self.assertEqual(third["auto_replaced_count"], 1)
            self.assertEqual(third["auto_error_count"], 1)
            self.assertEqual(third["error_count"], 2)

            rolled = store.rollback_latest_error_batch(
                rolled_back_at=now,
            )
            self.assertEqual(rolled["batch_id"], "manual-one")
            self.assertFalse(store.has_active_error_batches())
            connection = store._connect()
            try:
                counts = connection.execute(
                    "SELECT total_count, error_count FROM service_order_metrics"
                ).fetchone()
                active = connection.execute(
                    """
                    SELECT b.batch_type, d.order_number
                    FROM error_approval_details AS d
                    JOIN error_approval_batches AS b ON b.batch_id = d.batch_id
                    WHERE b.rolled_back_at IS NULL
                    """
                ).fetchall()
            finally:
                connection.close()
            self.assertEqual((counts["total_count"], counts["error_count"]), (3, 1))
            self.assertEqual([(row["batch_type"], row["order_number"]) for row in active], [("auto", "C")])

    def test_existing_batches_migrate_to_manual(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "legacy.sqlite3"
            connection = sqlite3.connect(path)
            try:
                connection.executescript(
                    """
                    CREATE TABLE error_approval_batches (
                        batch_id TEXT PRIMARY KEY, job_id TEXT NOT NULL,
                        source_name TEXT NOT NULL, approved_at TEXT NOT NULL,
                        data_start TEXT NOT NULL, data_end TEXT NOT NULL,
                        row_count INTEGER NOT NULL, rolled_back_at TEXT
                    );
                    CREATE TABLE error_approval_details (
                        detail_id INTEGER PRIMARY KEY AUTOINCREMENT,
                        batch_id TEXT NOT NULL, job_id TEXT NOT NULL,
                        candidate_row_id INTEGER NOT NULL, order_number TEXT NOT NULL,
                        order_date TEXT NOT NULL, person TEXT NOT NULL,
                        business TEXT NOT NULL, subcategory TEXT NOT NULL,
                        payload_json TEXT NOT NULL,
                        UNIQUE (batch_id, candidate_row_id)
                    );
                    INSERT INTO error_approval_batches VALUES
                        ('legacy', 'job', 'old.xlsx', '2026-07-01T00:00:00',
                         '2026-07-01', '2026-07-01', 0, NULL);
                    """
                )
                connection.commit()
            finally:
                connection.close()
            DashboardStore(path)
            connection = sqlite3.connect(path)
            try:
                batch_type = connection.execute(
                    "SELECT batch_type FROM error_approval_batches"
                ).fetchone()[0]
            finally:
                connection.close()
            self.assertEqual(batch_type, "manual")


if __name__ == "__main__":
    main()
